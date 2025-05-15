import tkinter as tk
from tkinter import ttk, filedialog
from typing import Dict, List, Tuple
from datetime import datetime, timedelta
from tkinter import simpledialog

import openpyxl
import xlrd
from customtkinter import CTk, CTkLabel, CTkEntry, CTkButton, CTkFrame, CTkScrollbar, CTkRadioButton, CTkComboBox
from tkinter import messagebox
from Item import NumberItem
from ItemManager import ItemManager
from GroupSolver import GroupSolver
from Users import check_credentials
from database import (create_database, add_order_to_db, delete_order_from_db, update_order_in_db,
                     get_all_orders_from_db, add_production_order, get_production_orders,
                     update_production_order_status, delete_production_order,
                     add_window_to_production_order, get_windows_for_production_order,
                     delete_window_from_production_order)

class AuthWindow(CTk):
    def __init__(self, parent):
        super().__init__()
        self.parent = parent
        self.geometry("300x200")
        self.title("Авторизация")
        self.protocol("WM_DELETE_WINDOW", self.on_close)  # Обрабатываем закрытие окна

        self.label_login = CTkLabel(self, text="Логин:")
        self.label_login.pack(pady=5)

        self.entry_login = CTkEntry(self, width=200)
        self.entry_login.pack(pady=5)

        self.label_password = CTkLabel(self, text="Пароль:")
        self.label_password.pack(pady=5)

        self.entry_password = CTkEntry(self, width=200, show="*")
        self.entry_password.pack(pady=5)
        self.entry_password.bind("<Return>", lambda event: self.authenticate())  # Обработка Enter в поле пароля

        self.button_login = CTkButton(self, text="Войти", command=self.authenticate)
        self.button_login.pack(pady=10)
        self.entry_login.bind("<Return>", lambda event: self.entry_password.focus())

    def authenticate(self):
        username = self.entry_login.get()
        password = self.entry_password.get()

        result = check_credentials(username, password)
        if result is True:
            self.destroy()
            self.parent.deiconify()  # Показываем главное окно
        elif result == "wrong_password":
            messagebox.showerror("Ошибка", "Неверный пароль.")
        elif result == "no_user":
            messagebox.showerror("Ошибка", "Пользователь не найден.")

    def on_close(self):
        if messagebox.askyesno("Выход", "Вы уверены, что хотите выйти?"):
            self.destroy()
            self.parent.destroy()  # Закрываем главное окно, чтобы программа завершилась корректно


class FrameCuttingTab(CTkFrame):
    def __init__(self, parent):
        super().__init__(parent)
        self.parent = parent
        self.groups = []  # Список для хранения карт раскроя

        # Левый фрейм для ввода заказов
        self.left_frame = CTkFrame(self, width=200)
        self.left_frame.pack(side=tk.LEFT, fill=tk.Y, padx=10, pady=10, expand=False)

        self.label_width = CTkLabel(self.left_frame, text="Ширина стеклопакета (мм):")
        self.label_width.pack(pady=5)
        self.entry_width = CTkEntry(self.left_frame, width=100)
        self.entry_width.pack(pady=5)

        self.label_height = CTkLabel(self.left_frame, text="Высота стеклопакета (мм):")
        self.label_height.pack(pady=5)
        self.entry_height = CTkEntry(self.left_frame, width=100)
        self.entry_height.pack(pady=5)

        self.label_type = CTkLabel(self.left_frame, text="Тип стеклопакета:")
        self.label_type.pack(pady=5)

        self.package_type = tk.StringVar(value="Однокамерный")

        self.radio_single = CTkRadioButton(self.left_frame, text="Однокамерный", variable=self.package_type,
                                           value="Однокамерный")
        self.radio_single.pack(anchor='w', padx=10)
        self.radio_double = CTkRadioButton(self.left_frame, text="Двухкамерный", variable=self.package_type,
                                           value="Двухкамерный")
        self.radio_double.pack(anchor='w', padx=10)

        self.add_order_button = CTkButton(self.left_frame, text="Добавить заказ", command=self.add_order)
        self.add_order_button.pack(pady=10)

        self.delete_order_button = CTkButton(self.left_frame, text="Удалить заказ", command=self.delete_order)
        self.delete_order_button.pack(pady=10)

        self.update_order_button = CTkButton(self.left_frame, text="Изменить заказ", command=self.update_order)
        self.update_order_button.pack(pady=10)

        self.order_list_frame = CTkFrame(self.left_frame)
        self.order_list_frame.pack(fill=tk.BOTH, expand=True, pady=10)

        self.scrollbar = CTkScrollbar(self.order_list_frame)
        self.scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        self.order_listbox = tk.Listbox(self.order_list_frame, yscrollcommand=self.scrollbar.set, height=15,
                                        bg="#333333", fg="white", font=("Arial", 12))
        self.order_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.scrollbar.configure(command=self.order_listbox.yview)

        # Правый фрейм для результатов оптимизации
        self.right_frame = CTkFrame(self, width=200)
        self.right_frame.pack(side=tk.RIGHT, fill=tk.Y, padx=10, pady=10, expand=False)

        self.card_list_frame = CTkFrame(self.right_frame)
        self.card_list_frame.pack(fill=tk.BOTH, expand=True, pady=10)

        self.card_listbox = tk.Listbox(self.card_list_frame, height=15, bg="#333333", fg="white", font=("Arial", 12))
        self.card_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        self.scrollbar_cards = CTkScrollbar(self.card_list_frame)
        self.scrollbar_cards.pack(side=tk.RIGHT, fill=tk.Y)
        self.scrollbar_cards.configure(command=self.card_listbox.yview)

        # Центральная область
        self.center_frame = CTkFrame(self)
        self.center_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=10, pady=10)

        self.optimize_button = CTkButton(self.center_frame, text="Запустить оптимизацию", command=self.optimize_cutting,
                                         width=200)
        self.optimize_button.pack(pady=5, anchor='n')

        self.card_canvas = tk.Canvas(self.center_frame, width=800, height=100, bg="white")
        self.card_canvas.pack(pady=10)

        self.unused_label = CTkLabel(self.center_frame, text=f"")
        self.unused_label.pack(anchor='w', padx=10, pady=5)

        # Загружаем заказы из базы данных
        self.load_orders_from_db()

    def add_order(self):
        width = self.entry_width.get()
        height = self.entry_height.get()
        if width and height:
            try:
                width = int(width)
                height = int(height)
                if width <= 0 or height <= 0:
                    messagebox.showerror("Ошибка", "Ширина и высота должны быть положительными числами.")
                    return

                # Добавляем заказ в базу данных
                add_order_to_db(width, height, self.package_type.get())

                # Обновляем список заказов
                self.load_orders_from_db()

                # Оповещаем родительское окно об изменении данных
                self.parent.on_orders_updated()

            except ValueError:
                messagebox.showerror("Ошибка", "Неверный ввод. Пожалуйста, введите целые числа.")

    def delete_order(self):
        selected_order_index = self.order_listbox.curselection()
        if selected_order_index:
            # Получаем ID выбранного заказа
            orders = get_all_orders_from_db()
            order_id = orders[selected_order_index[0]][0]

            delete_order_from_db(order_id)
            self.load_orders_from_db()
            self.parent.on_orders_updated()
        else:
            messagebox.showwarning("Предупреждение", "Пожалуйста, выберите заказ для удаления.")

    def update_order(self):
        selected_order_index = self.order_listbox.curselection()
        if selected_order_index:
            # Получаем выбранный заказ
            orders = get_all_orders_from_db()
            order_id = orders[selected_order_index[0]][0]

            new_width = self.entry_width.get()
            new_height = self.entry_height.get()

            if new_width and new_height:
                try:
                    new_width = int(new_width)
                    new_height = int(new_height)

                    update_order_in_db(order_id, new_width, new_height)
                    self.load_orders_from_db()
                    self.parent.on_orders_updated()

                    self.entry_width.delete(0, tk.END)
                    self.entry_height.delete(0, tk.END)

                except ValueError:
                    messagebox.showerror("Ошибка", "Неверный ввод. Пожалуйста, введите целые числа.")
            else:
                messagebox.showwarning("Предупреждение", "Пожалуйста, введите новые значения.")
        else:
            messagebox.showwarning("Предупреждение", "Пожалуйста, выберите заказ для изменения.")

    def load_orders_from_db(self):
        orders = get_all_orders_from_db()
        self.order_listbox.delete(0, tk.END)
        for order in orders:
            order_id, width, height, package_type = order
            order_text = f"Заказ {order_id}: {width}x{height} ({package_type})"
            self.order_listbox.insert(tk.END, order_text)

    def optimize_cutting(self):
        orders = get_all_orders_from_db()
        items = []

        for order in orders:
            order_id, width, height, package_type = order
            if package_type == "Однокамерный":
                values = [width, height, width, height]
            else:
                values = [width, height, width, height, width, height, width, height]

            items.append(NumberItem(index=order_id, values=values))

        solver = GroupSolver(items)
        groups, unused_values = solver.group_numbers()

        self.groups = groups
        self.card_listbox.delete(0, tk.END)

        for idx, group in enumerate(groups):
            group_sum = sum(value for values in group.values() for value in values)
            group_text = f"Карта {idx + 1}: Сумма = {group_sum}"
            self.card_listbox.insert(tk.END, group_text)

        if unused_values:
            self.unused_label.configure(text=f"Не задействованные значения: {unused_values}")

    def draw_horizontal_cutting_plan(self, group):
        canvas_width = 800
        canvas_height = 100
        self.card_canvas.delete("all")

        total_length = 6000
        x_start = 0
        used_length = 0

        for order_id, lengths in group.items():
            for length in lengths:
                rect_width = (length / total_length) * canvas_width
                x_end = x_start + rect_width

                y_top = (canvas_height - 30) / 2
                y_bottom = (canvas_height + 30) / 2
                self.card_canvas.create_rectangle(
                    x_start, y_top, x_end, y_bottom, fill="green", outline="black"
                )

                self.card_canvas.create_text(
                    (x_start + x_end) / 2, (y_top + y_bottom) / 2,
                    text=f"{length}", fill="black", font=("Arial", 10)
                )

                self.card_canvas.create_text(
                    (x_start + x_end) / 2, canvas_height - 10,
                    text=f"{order_id}", fill="blue", font=("Arial", 10)
                )

                x_start = x_end
                used_length += length

        remaining_length = total_length - used_length
        if remaining_length > 0:
            rect_width = max((remaining_length / total_length) * canvas_width, 1)
            x_end = x_start + rect_width

            y_top = (canvas_height - 30) / 2
            y_bottom = (canvas_height + 30) / 2

            self.card_canvas.create_rectangle(
                x_start, y_top, x_end, y_bottom, fill="red", outline="black"
            )

            self.card_canvas.create_text(
                (x_start + x_end) / 2, (y_top + y_bottom) / 2,
                text=f"{remaining_length} мм", fill="white", font=("Arial", 10)
            )

    def display_card_details(self, event):
        selected_index = self.card_listbox.curselection()
        if selected_index:
            group = self.groups[selected_index[0]]
            self.draw_horizontal_cutting_plan(group)


class GlassCuttingTab(CTkFrame):
    def __init__(self, parent):
        super().__init__(parent)
        self.cutting_margin = 5  # Технологический зазор между деталями (мм)
        self.min_rotation_angle = 90  # Минимальный шаг поворота
        self.parent = parent
        self.groups = []
        self.sheet_width = 6000  # Ширина листа стекла по умолчанию
        self.sheet_height = 6000  # Высота листа стекла по умолчанию
        self.zoom_level = 0.8

        self.selected_item = None
        self.hover_item = None
        self.selection_rect = None
        self.hover_rect = None
        self.tooltip = None


        # Левый фрейм для работы с заказами
        self.left_frame = CTkFrame(self, width=200)
        self.left_frame.pack(side=tk.LEFT, fill=tk.Y, padx=10, pady=10, expand=False)

        # Добавляем поля для ввода размеров листа стекла
        self.label_sheet_width = CTkLabel(self.left_frame, text="Ширина листа стекла (мм):")
        self.label_sheet_width.pack(pady=5)
        self.entry_sheet_width = CTkEntry(self.left_frame, width=100)
        self.entry_sheet_width.insert(0, "6000")
        self.entry_sheet_width.pack(pady=5)

        self.label_sheet_height = CTkLabel(self.left_frame, text="Высота листа стекла (мм):")
        self.label_sheet_height.pack(pady=5)
        self.entry_sheet_height = CTkEntry(self.left_frame, width=100)
        self.entry_sheet_height.insert(0, "6000")
        self.entry_sheet_height.pack(pady=5)

        self.label_width = CTkLabel(self.left_frame, text="Ширина стеклопакета (мм):")
        self.label_width.pack(pady=5)
        self.entry_width = CTkEntry(self.left_frame, width=100)
        self.entry_width.pack(pady=5)

        self.label_height = CTkLabel(self.left_frame, text="Высота стеклопакета (мм):")
        self.label_height.pack(pady=5)
        self.entry_height = CTkEntry(self.left_frame, width=100)
        self.entry_height.pack(pady=5)

        self.label_type = CTkLabel(self.left_frame, text="Тип стеклопакета:")
        self.label_type.pack(pady=5)

        self.package_type = tk.StringVar(value="Однокамерный")

        self.radio_single = CTkRadioButton(self.left_frame, text="Однокамерный", variable=self.package_type,
                                       value="Однокамерный")
        self.radio_single.pack(anchor='w', padx=10)
        self.radio_double = CTkRadioButton(self.left_frame, text="Двухкамерный", variable=self.package_type,
                                       value="Двухкамерный")
        self.radio_double.pack(anchor='w', padx=10)

        self.add_order_button = CTkButton(self.left_frame, text="Добавить заказ", command=self.add_order)
        self.add_order_button.pack(pady=10)

        self.delete_order_button = CTkButton(self.left_frame, text="Удалить заказ", command=self.delete_order)
        self.delete_order_button.pack(pady=10)

        self.update_order_button = CTkButton(self.left_frame, text="Изменить заказ", command=self.update_order)
        self.update_order_button.pack(pady=10)

        self.order_list_frame = CTkFrame(self.left_frame)
        self.order_list_frame.pack(fill=tk.BOTH, expand=True, pady=10)

        self.scrollbar = CTkScrollbar(self.order_list_frame)
        self.scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        self.order_listbox = tk.Listbox(self.order_list_frame, yscrollcommand=self.scrollbar.set, height=15,
                                    bg="#333333", fg="white", font=("Arial", 12))
        self.order_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.scrollbar.configure(command=self.order_listbox.yview)

        # Правый фрейм для результатов
        self.right_frame = CTkFrame(self, width=200)
        self.right_frame.pack(side=tk.RIGHT, fill=tk.Y, padx=10, pady=10, expand=False)

        self.card_list_frame = CTkFrame(self.right_frame)
        self.card_list_frame.pack(fill=tk.BOTH, expand=True, pady=10)

        self.card_listbox = tk.Listbox(self.card_list_frame, height=15, bg="#333333", fg="white", font=("Arial", 12))
        self.card_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        self.scrollbar_cards = CTkScrollbar(self.card_list_frame)
        self.scrollbar_cards.pack(side=tk.RIGHT, fill=tk.Y)
        self.scrollbar_cards.configure(command=self.card_listbox.yview)

        # Центральная область
        self.center_frame = CTkFrame(self)
        self.center_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Кнопка оптимизации сверху
        self.optimize_button = CTkButton(
            self.center_frame,
            text="Запустить оптимизацию",
            command=self.optimize_cutting,
            width=200
        )
        self.optimize_button.pack(pady=5, anchor='n')

        # Холст для отображения (как было раньше)
        self.card_canvas = tk.Canvas(
            self.center_frame,
            width=600,
            height=600,
            bg="white",
            highlightthickness=0
        )
        self.card_canvas.pack(pady=10, fill=tk.BOTH, expand=True)

        # Метка для неиспользованной области (как было)
        self.unused_label = CTkLabel(self.center_frame, text="")
        self.unused_label.pack(anchor='w', padx=10, pady=5)
        # Привязка событий для нового функционала
        self.card_canvas.bind("<Motion>", self.on_canvas_hover)
        self.card_canvas.bind("<Button-1>", self.on_canvas_click)
        self.card_canvas.bind("<Leave>", self.hide_tooltip)

        # Привязываем обработчик выбора карты
        self.card_listbox.bind('<<ListboxSelect>>', self.on_card_select)
        self.select_default_card()

        self.load_orders_from_db()

    def select_default_card(self):
        """Выбирает первую карту раскроя по умолчанию"""
        if self.card_listbox.size() > 0:
            self.card_listbox.selection_set(0)
            self.card_listbox.see(0)
            self.display_cutting_plan(0)

    def on_canvas_hover(self, event):
        """Всплывающая подсказка при наведении"""
        if not self.groups or not self.card_listbox.curselection():
            self.hide_tooltip()
            return

        group = self.groups[self.card_listbox.curselection()[0]]
        scale = self.get_current_scale(group)

        # Ищем заготовку под курсором
        new_hover_item = None
        for item in group['items']:
            x1 = item['x'] * scale
            y1 = item['y'] * scale
            x2 = x1 + item['width'] * scale
            y2 = y1 + item['height'] * scale

            if x1 <= event.x <= x2 and y1 <= event.y <= y2:
                new_hover_item = item
                break

        # Обновляем hover-эффект только если курсор перешел на новый элемент или вышел с элемента
        if new_hover_item != self.hover_item:
            self.update_hover_effect(new_hover_item, scale)
            self.hover_item = new_hover_item

            # Показываем tooltip только если нашли элемент
            if new_hover_item:
                self.show_tooltip(event.x, event.y, new_hover_item)
            else:
                self.hide_tooltip()  # Скрываем подсказку если курсор на пустом месте

        # Если курсор перемещается по тому же элементу, просто обновляем позицию tooltip
        elif new_hover_item and hasattr(self, 'tooltip_bg') and self.tooltip_bg:
            # Перемещаем существующую подсказку
            self.card_canvas.coords(self.tooltip_bg, event.x + 10, event.y + 10, event.x + 150, event.y + 40)
            self.card_canvas.coords(self.tooltip_text, event.x + 15, event.y + 15)

        # Дополнительная проверка: если курсор на пустом месте и есть подсказка - скрываем
        elif not new_hover_item and (hasattr(self, 'tooltip_bg') and self.tooltip_bg):
            self.hide_tooltip()

    def update_hover_effect(self, item, scale):
        """Обновляет подсветку при наведении"""
        if self.hover_rect:
            self.card_canvas.delete(self.hover_rect)
            self.hover_rect = None

        if item:
            x1 = item['x'] * scale
            y1 = item['y'] * scale
            x2 = x1 + item['width'] * scale
            y2 = y1 + item['height'] * scale

            self.hover_rect = self.card_canvas.create_rectangle(
                x1, y1, x2, y2,
                outline="#FFA500", width=2, dash=(3, 3)
            )
            self.card_canvas.tag_raise(self.hover_rect)

    def show_tooltip(self, x, y, item):
        """Показывает всплывающую подсказку"""
        # Сначала скрываем предыдущую подсказку
        self.hide_tooltip()

        text = f"ID: {item['id']} | {item['width']}×{item['height']} мм"
        if item.get('rotation'):
            text += f" (повернуто)"

        # Создаем фон для tooltip сначала (чтобы текст был поверх него)
        self.tooltip_bg = self.card_canvas.create_rectangle(
            x + 10, y + 10,
            x + 150, y + 40,  # Фиксированный размер, чтобы не вычислять bbox
            fill="#FFFFE0",
            outline="#CCCCCC",
            tags="tooltip"
        )

        # Затем создаем текст
        self.tooltip_text = self.card_canvas.create_text(
            x + 15, y + 15,
            text=text,
            font=("Arial", 10),
            fill="black",
            anchor="nw",
            tags="tooltip"
        )

    def hide_tooltip(self, event=None):
        """Скрывает всплывающую подсказку"""
        if hasattr(self, 'tooltip_bg') and self.tooltip_bg:
            self.card_canvas.delete(self.tooltip_bg)
            self.tooltip_bg = None

        if hasattr(self, 'tooltip_text') and self.tooltip_text:
            self.card_canvas.delete(self.tooltip_text)
            self.tooltip_text = None

        if hasattr(self, 'hover_rect') and self.hover_rect:
            self.card_canvas.delete(self.hover_rect)
            self.hover_rect = None

        self.hover_item = None

    def on_canvas_click(self, event):
        """Упрощенный обработчик клика с гарантией выбранной карты"""
        try:
            # Гарантируем, что есть выбранная карта
            if not self.groups or not self.card_listbox.curselection():
                self.card_listbox.selection_set(0)  # Форсируем выбор первой карты

            group = self.groups[self.card_listbox.curselection()[0]]
            scale = self.get_current_scale(group)

            # Остальная логика обработки клика остается без изменений
            self.clear_selection()

            clicked_item = None
            for item in group['items']:
                x1 = item['x'] * scale
                y1 = item['y'] * scale
                x2 = x1 + item['width'] * scale
                y2 = y1 + item['height'] * scale

                if x1 <= event.x <= x2 and y1 <= event.y <= y2:
                    clicked_item = item
                    break

            if clicked_item:
                self.create_selection_rect(clicked_item, scale)
                self.selected_item = clicked_item
                self.select_order_in_list(clicked_item['id'])

        except Exception as e:
            print(f"Ошибка: {e}")
            self.clear_selection()

    def create_selection_rect(self, item, scale):
        """Создает прямоугольник выделения"""
        x1 = item['x'] * scale
        y1 = item['y'] * scale
        x2 = x1 + item['width'] * scale
        y2 = y1 + item['height'] * scale

        self.selection_rect = self.card_canvas.create_rectangle(
            x1, y1, x2, y2,
            outline="#00FF00", width=3, tags="selection"
        )
        self.card_canvas.tag_raise(self.selection_rect)

    def clear_selection(self):
        """Сбрасывает выделение"""
        if hasattr(self, 'selection_rect') and self.selection_rect:
            self.card_canvas.delete(self.selection_rect)
        self.selection_rect = None
        self.selected_item = None
        if hasattr(self, 'order_listbox'):
            self.order_listbox.selection_clear(0, tk.END)

    def on_card_select(self, event):
        """Обработчик выбора карты раскроя"""
        if not self.card_listbox.curselection():
            return

        selected_index = self.card_listbox.curselection()[0]
        if 0 <= selected_index < len(self.groups):
            self.display_cutting_plan(selected_index)
            self.clear_selection()  # Сбрасываем выделение при смене карты

    def select_item(self, item, scale):
        """Выделяет указанную заготовку (без предварительной очистки)"""
        x1 = item['x'] * scale
        y1 = item['y'] * scale
        x2 = x1 + item['width'] * scale
        y2 = y1 + item['height'] * scale

        self.selection_rect = self.card_canvas.create_rectangle(
            x1, y1, x2, y2,
            outline="#00FF00", width=3, tags="selection"
        )
        self.card_canvas.tag_raise(self.selection_rect)
        self.selected_item = item


    def select_order_in_list(self, order_id):
        """Выбирает заказ в списке с защитой от ошибок"""
        try:
            for i in range(self.order_listbox.size()):
                if str(order_id) in self.order_listbox.get(i):
                    self.order_listbox.selection_clear(0, tk.END)
                    self.order_listbox.selection_set(i)
                    self.order_listbox.see(i)
                    break
        except Exception as e:
            print(f"Ошибка при выборе заказа: {e}")

    def get_current_scale(self, group):
        """Возвращает текущий масштаб отображения"""
        return min(
            self.card_canvas.winfo_width() / group['width'],
            self.card_canvas.winfo_height() / group['height']
        )

    def add_order(self):
        width = self.entry_width.get()
        height = self.entry_height.get()
        if width and height:
            try:
                width = int(width)
                height = int(height)
                if width <= 0 or height <= 0:
                    messagebox.showerror("Ошибка", "Ширина и высота должны быть положительными числами.")
                    return

                add_order_to_db(width, height, self.package_type.get())
                self.load_orders_from_db()
                self.parent.on_orders_updated()

            except ValueError:
                messagebox.showerror("Ошибка", "Неверный ввод. Пожалуйста, введите целые числа.")

    def delete_order(self):
        selected_order_index = self.order_listbox.curselection()
        if selected_order_index:
            orders = get_all_orders_from_db()
            order_id = orders[selected_order_index[0]][0]

            delete_order_from_db(order_id)
            self.load_orders_from_db()
            self.parent.on_orders_updated()
        else:
            messagebox.showwarning("Предупреждение", "Пожалуйста, выберите заказ для удаления.")

    def update_order(self):
        selected_order_index = self.order_listbox.curselection()
        if selected_order_index:
            orders = get_all_orders_from_db()
            order_id = orders[selected_order_index[0]][0]

            new_width = self.entry_width.get()
            new_height = self.entry_height.get()

            if new_width and new_height:
                try:
                    new_width = int(new_width)
                    new_height = int(new_height)

                    update_order_in_db(order_id, new_width, new_height)
                    self.load_orders_from_db()
                    self.parent.on_orders_updated()

                    self.entry_width.delete(0, tk.END)
                    self.entry_height.delete(0, tk.END)

                except ValueError:
                    messagebox.showerror("Ошибка", "Неверный ввод. Пожалуйста, введите целые числа.")
            else:
                messagebox.showwarning("Предупреждение", "Пожалуйста, введите новые значения.")
        else:
            messagebox.showwarning("Предупреждение", "Пожалуйста, выберите заказ для изменения.")

    def load_orders_from_db(self):
        orders = get_all_orders_from_db()
        self.order_listbox.delete(0, tk.END)
        for order in orders:
            order_id, width, height, package_type = order
            order_text = f"Заказ {order_id}: {width}x{height} ({package_type})"
            self.order_listbox.insert(tk.END, order_text)

    def optimize_cutting(self):
        """Усовершенствованный алгоритм оптимизации раскроя с поворотами"""
        try:
            self.sheet_width = int(self.entry_sheet_width.get())
            self.sheet_height = int(self.entry_sheet_height.get())
        except ValueError:
            messagebox.showerror("Ошибка", "Неверные размеры листа стекла")
            return

        orders = get_all_orders_from_db()
        if not orders:
            messagebox.showwarning("Предупреждение", "Нет заказов для оптимизации.")
            return

        # Подготовка данных
        items = [{'id': o[0], 'width': o[1], 'height': o[2]} for o in orders]

        # Алгоритм Best-Fit Decreasing с поворотами
        self.groups = self.best_fit_decreasing_algorithm(items)

        # Обновляем интерфейс
        self.update_interface()

        self.select_default_card()  # Автовыбор первой карты

    def update_interface(self):
        """Обновляет все элементы интерфейса после оптимизации"""
        self.card_listbox.delete(0, tk.END)

        for i, group in enumerate(self.groups):
            used_area = sum(i['width'] * i['height'] for i in group['items'])
            total_area = group['width'] * group['height']
            utilization = used_area / total_area * 100
            self.card_listbox.insert(tk.END, f"Карта {i + 1} (Использовано: {utilization:.1f}%)")

        if self.groups:
            self.display_cutting_plan(0)

    def best_fit_decreasing_algorithm(self, items: List[Dict]) -> List[Dict]:
        """Алгоритм Best-Fit Decreasing с поддержкой поворотов"""
        # Сортируем элементы по убыванию площади
        sorted_items = sorted(items, key=lambda x: x['width'] * x['height'], reverse=True)

        sheets = []

        for item in sorted_items:
            placed = False

            # Пробуем разместить на существующих листах
            for sheet in sheets:
                result = self.try_place_item(sheet, item)
                if result:
                    placed = True
                    break

            # Если не поместилось, создаем новый лист
            if not placed:
                new_sheet = {
                    'width': self.sheet_width,
                    'height': self.sheet_height,
                    'items': [],
                    'remaining_rectangles': [(0, 0, self.sheet_width, self.sheet_height)]
                }
                self.try_place_item(new_sheet, item)
                sheets.append(new_sheet)

        return sheets

    def try_place_item(self, sheet: Dict, item: Dict) -> bool:
        """Пытается разместить элемент на листе с оптимальным поворотом"""
        best_rotation = None
        best_fit = None
        best_rect = None
        min_waste = float('inf')

        # Проверяем все возможные повороты (0, 90 градусов)
        for rotation in [0, 90]:
            if rotation == 0:
                w, h = item['width'], item['height']
            else:
                w, h = item['height'], item['width']

            # Ищем лучшее место для размещения
            for rect in sheet['remaining_rectangles']:
                rx, ry, rw, rh = rect

                # Проверяем, помещается ли элемент
                if w <= rw and h <= rh:
                    # Вычисляем "отходы" (оставшееся пространство)
                    waste = (rw * rh) - (w * h)

                    # Если нашли лучшее место
                    if waste < min_waste:
                        min_waste = waste
                        best_rotation = rotation
                        best_fit = (w, h)
                        best_rect = rect

        # Если нашли подходящее место
        if best_rect:
            rx, ry, rw, rh = best_rect
            w, h = best_fit

            # Добавляем элемент
            sheet['items'].append({
                'id': item['id'],
                'x': rx,
                'y': ry,
                'width': w,
                'height': h,
                'rotation': best_rotation
            })

            # Обновляем оставшееся пространство (алгоритм Guillotine)
            self.update_remaining_space(sheet, best_rect, w, h)
            return True

        return False

    def update_remaining_space(self, sheet: Dict, rect: Tuple, w: int, h: int):
        """Обновляет оставшееся пространство по алгоритму Guillotine"""
        rx, ry, rw, rh = rect
        sheet['remaining_rectangles'].remove(rect)

        # Оставшееся пространство справа
        if rw - w > 0:
            sheet['remaining_rectangles'].append((rx + w, ry, rw - w, h))

        # Оставшееся пространство сверху
        if rh - h > 0:
            sheet['remaining_rectangles'].append((rx, ry + h, w, rh - h))

        # Сортировка оставшихся прямоугольников по площади
        sheet['remaining_rectangles'].sort(key=lambda r: r[2] * r[3], reverse=True)

    def display_cutting_plan(self, index):
        """Отображение карты раскроя с сеткой и неиспользованными областями"""
        if not self.groups or index >= len(self.groups):
            return

        group = self.groups[index]
        self.card_canvas.delete("all")

        # Масштабирование
        canvas_width = self.card_canvas.winfo_width()
        canvas_height = self.card_canvas.winfo_height()
        scale = min(canvas_width / group['width'], canvas_height / group['height'])

        # 1. Рисуем фон (красный горошек для всего листа)
        self.card_canvas.create_rectangle(
            0, 0,
            group['width'] * scale,
            group['height'] * scale,
            fill="red", stipple="gray25", outline=""
        )

        # 2. Рисуем сетку с шагом 1000 мм
        self.draw_grid(group, scale)

        # 3. Рисуем границы листа поверх сетки
        self.card_canvas.create_rectangle(
            0, 0,
            group['width'] * scale,
            group['height'] * scale,
            outline="black", width=3
        )

        # 4. Рисуем все элементы поверх всего
        for item in group['items']:
            self.draw_glass_item(item, scale)

        # 5. Информационная панель
        self.update_info_panel(group, index, scale)

        # Восстанавливаем выделение после перерисовки
        if hasattr(self, 'selected_item') and self.selected_item:
            for item in group['items']:
                if item['id'] == self.selected_item['id']:
                    self.select_item(item, scale)
                    break

    def draw_grid(self, group, scale):
        """Рисует сетку с шагом 1000 мм"""
        grid_color = "#cccccc"
        grid_step = 1000  # Шаг сетки в мм

        # Вертикальные линии
        for x in range(0, group['width'] + grid_step, grid_step):
            x_pos = x * scale
            self.card_canvas.create_line(
                x_pos, 0,
                x_pos, group['height'] * scale,
                fill=grid_color, dash=(2, 2)
            )
            # Подписи осей X
            if x > 0:
                self.card_canvas.create_text(
                    x_pos, 10,
                    text=f"{x} мм",
                    font=("Arial", 8),
                    anchor=tk.N
                )

        # Горизонтальные линии
        for y in range(0, group['height'] + grid_step, grid_step):
            y_pos = y * scale
            self.card_canvas.create_line(
                0, y_pos,
                group['width'] * scale, y_pos,
                fill=grid_color, dash=(2, 2)
            )
            # Подписи осей Y
            if y > 0:
                self.card_canvas.create_text(
                    10, y_pos,
                    text=f"{y} мм",
                    font=("Arial", 8),
                    anchor=tk.W
                )

    def draw_glass_item(self, item, scale):
        """Рисует одну заготовку с подписью"""
        x1 = item['x'] * scale
        y1 = item['y'] * scale
        x2 = x1 + item['width'] * scale
        y2 = y1 + item['height'] * scale

        # Цвет в зависимости от поворота
        color = "#4CAF50" if item['rotation'] == 0 else "#2196F3"

        # Рисуем элемент
        self.card_canvas.create_rectangle(
            x1, y1, x2, y2,
            outline="black", fill=color, width=1
        )

        # Подпись (если элемент достаточно большой)
        if (x2 - x1) > 60 and (y2 - y1) > 40:
            text = f"{item['width']}×{item['height']}\nID:{item['id']}"
            if item['rotation'] != 0:
                text += f"\n(повернуто)"

            font_size = max(8, min(12, int(min(x2 - x1, y2 - y1) / 15)))
            self.card_canvas.create_text(
                (x1 + x2) / 2, (y1 + y2) / 2,
                text=text, font=("Arial", font_size),
                fill="black"
            )

    def update_info_panel(self, group, index, scale):
        """Обновляет информационную панель"""
        used_area = sum(i['width'] * i['height'] for i in group['items'])
        total_area = group['width'] * group['height']
        utilization = used_area / total_area * 100

        self.unused_label.configure(
            text=f"Карта {index + 1}/{len(self.groups)} | "
                 f"Размер: {group['width']}×{group['height']} мм | "
                 f"Использовано: {utilization:.1f}% | "
                 f"Масштаб: 1:{int(1 / scale)} | "
                 f"Сетка: 1000 мм"
        )
    #
    # def calculate_utilization(self, group):
    #     """Вычисляет процент использования листа"""
    #     used_area = sum(item['width'] * item['height'] for item in group['items'])
    #     total_area = group['width'] * group['height']
    #     return round(used_area / total_area * 100, 1)

    def display_card_details(self, event):
        selected_index = self.card_listbox.curselection()
        if selected_index:
            self.display_cutting_plan(selected_index[0])


class ProductionPlanningTab(CTkFrame):
    def __init__(self, parent):
        super().__init__(parent)
        self.parent = parent
        self.current_order_id = None
        self._first_draw = True  # Флаг первого отображения

        # Основные фреймы
        self.left_frame = CTkFrame(self, width=300)
        self.left_frame.pack(side=tk.LEFT, fill=tk.Y, padx=10, pady=10)

        self.right_frame = CTkFrame(self)
        self.right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Форма добавления заказа
        self.add_order_frame = CTkFrame(self.left_frame)
        self.add_order_frame.pack(fill=tk.X, pady=5)

        CTkLabel(self.add_order_frame, text="Новый производственный заказ").pack(pady=5)

        # Поля ввода
        self.order_name_entry = CTkEntry(self.add_order_frame, placeholder_text="Название заказа")
        self.order_name_entry.pack(fill=tk.X, pady=5)

        self.customer_entry = CTkEntry(self.add_order_frame, placeholder_text="Заказчик")
        self.customer_entry.pack(fill=tk.X, pady=5)

        self.deadline_entry = CTkEntry(self.add_order_frame, placeholder_text="Срок (ДД.ММ.ГГГГ)")
        self.deadline_entry.pack(fill=tk.X, pady=5)

        self.priority_var = tk.StringVar(value="Средний")
        self.priority_combo = CTkComboBox(self.add_order_frame,
                                          values=["Низкий", "Средний", "Высокий", "Критичный"],
                                          variable=self.priority_var)
        self.priority_combo.pack(fill=tk.X, pady=5)

        self.add_button = CTkButton(self.add_order_frame, text="Добавить заказ", command=self.add_production_order)
        self.add_button.pack(fill=tk.X, pady=5)

        self.import_button = CTkButton(self.add_order_frame, text="Импорт заказа",
                                       command=self.import_order_from_excel)
        self.import_button.pack(fill=tk.X, pady=5)

        # Список заказов
        self.orders_frame = CTkFrame(self.left_frame)
        self.orders_frame.pack(fill=tk.BOTH, expand=True, pady=5)

        self.orders_listbox = tk.Listbox(self.orders_frame, bg="#333333", fg="white", font=("Arial", 12))
        self.orders_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.orders_listbox.bind('<<ListboxSelect>>', self.show_order_details)

        self.scrollbar = CTkScrollbar(self.orders_frame)
        self.scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.orders_listbox.config(yscrollcommand=self.scrollbar.set)
        self.scrollbar.configure(command=self.orders_listbox.yview)
        self.load_production_orders()

        # Кнопки управления
        self.control_frame = CTkFrame(self.left_frame)
        self.control_frame.pack(fill=tk.X, pady=5)

        self.start_button = CTkButton(self.control_frame, text="В работу",
                                      command=lambda: self.change_order_status("В работе"))
        self.start_button.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=2)

        self.complete_button = CTkButton(self.control_frame, text="Завершить",
                                         command=lambda: self.change_order_status("Завершен"))
        self.complete_button.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=2)

        self.delete_button = CTkButton(self.control_frame, text="Удалить", command=self.delete_order)
        self.delete_button.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=2)

        # Детали заказа
        self.details_frame = CTkFrame(self.right_frame)
        self.details_frame.pack(fill=tk.BOTH, expand=True, pady=5)

        # Вкладки для деталей заказа
        self.details_notebook = ttk.Notebook(self.details_frame)
        self.details_notebook.pack(fill=tk.BOTH, expand=True)

        # Вкладка с информацией о заказе
        self.info_tab = CTkFrame(self.details_notebook)
        self.details_notebook.add(self.info_tab, text="Информация")

        self.order_details_text = tk.Text(self.info_tab, bg="#333333", fg="white", font=("Arial", 12))
        self.order_details_text.pack(fill=tk.BOTH, expand=True)

        # Вкладка со стеклопакетами
        self.windows_tab = CTkFrame(self.details_notebook)
        self.details_notebook.add(self.windows_tab, text="Стеклопакеты")

        # Кнопки управления стеклопакетами
        self.windows_control_frame = CTkFrame(self.windows_tab)
        self.windows_control_frame.pack(fill=tk.X, pady=5)

        self.add_window_button = CTkButton(self.windows_control_frame, text="Добавить стеклопакет",
                                           command=self.add_window_to_order)
        self.add_window_button.pack(side=tk.LEFT, padx=5)

        self.delete_window_button = CTkButton(self.windows_control_frame, text="Удалить стеклопакет",
                                              command=self.delete_window_from_order)
        self.delete_window_button.pack(side=tk.LEFT, padx=5)

        # Таблица стеклопакетов
        self.windows_tree = ttk.Treeview(self.windows_tab, columns=("id", "type", "width", "height", "quantity"), show="headings")
        self.windows_tree.heading("id", text="ID")
        self.windows_tree.heading("type", text="Тип")
        self.windows_tree.heading("width", text="Ширина (мм)")
        self.windows_tree.heading("height", text="Высота (мм)")
        self.windows_tree.heading("quantity", text="Количество")
        self.windows_tree.column("id", width=50)
        self.windows_tree.column("type", width=150)
        self.windows_tree.column("width", width=75)
        self.windows_tree.column("height", width=75)
        self.windows_tree.column("quantity", width=50)

        scrollbar = ttk.Scrollbar(self.windows_tab, orient="vertical", command=self.windows_tree.yview)
        self.windows_tree.configure(yscrollcommand=scrollbar.set)

        self.windows_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # Календарь производства
        self.calendar_frame = CTkFrame(self.right_frame)
        self.calendar_frame.pack(fill=tk.BOTH, expand=True, pady=5)

       # Загрузка данных
        self.load_production_orders()

        # Загрузка данных и отрисовка календаря
        self.after(100, self.update_calendar)  # Запускаем через небольшой таймаут
        self.bind("<Visibility>", self.on_visibility_changed)
        self.bind("<Configure>", self.on_resize)  # Обработчик изменения размеров

        # Панель навигации по месяцам
        self.calendar_nav_frame = CTkFrame(self.calendar_frame)
        self.calendar_nav_frame.pack(fill=tk.X, pady=5)

        self.prev_month_btn = CTkButton(self.calendar_nav_frame, text="<", width=30,
                                        command=self.show_prev_month)
        self.prev_month_btn.pack(side=tk.LEFT, padx=5)

        self.month_label = CTkLabel(self.calendar_nav_frame, text="", font=("Arial", 12))
        self.month_label.pack(side=tk.LEFT, expand=True)

        self.next_month_btn = CTkButton(self.calendar_nav_frame, text=">", width=30,
                                        command=self.show_next_month)
        self.next_month_btn.pack(side=tk.RIGHT, padx=5)

        # Холст календаря
        self.calendar_canvas = tk.Canvas(self.calendar_frame, bg="white")
        self.calendar_canvas.pack(fill=tk.BOTH, expand=True)
        self.calendar_canvas.bind("<Button-1>", self.on_calendar_click)

        # Текущий месяц и год для отображения
        self.current_date = datetime.now().date()
        self.current_month = self.current_date.month
        self.current_year = self.current_date.year

        # Загрузка данных и отрисовка календаря
        self.update_calendar()

    def on_visibility_changed(self, event):
        """Обработчик изменения видимости вкладки"""
        if self._first_draw and self.winfo_ismapped():
            self._first_draw = False
            self.after(100, self.update_calendar)

    def on_resize(self, event):
        """Обработчик изменения размеров"""
        if self.winfo_ismapped():  # Проверяем, что вкладка видима
            self.update_calendar()

    def update_calendar(self):
        """Обновление отображения календаря"""
        if not self.winfo_ismapped():  # Не обновляем, если вкладка не видна
            return

        if not self.calendar_canvas.winfo_exists():
            return

        self.calendar_canvas.delete("all")
        self.month_label.configure(text=f"{self.current_month}.{self.current_year}")

        # Получаем все заказы
        orders = get_production_orders()

        # Определяем первый и последний день месяца
        first_day = datetime(self.current_year, self.current_month, 1).date()
        last_day = datetime(self.current_year, self.current_month + 1, 1).date() - timedelta(days=1)

        # Определяем дни недели для первого и последнего дня месяца
        start_weekday = first_day.weekday()  # 0-пн, 6-вс
        total_days = last_day.day

        # Настройки отображения
        canvas_width = self.calendar_canvas.winfo_width()
        canvas_height = self.calendar_canvas.winfo_height()
        if canvas_width < 10 or canvas_height < 10:  # Минимальные размеры
            return
        day_width = canvas_width / 7
        day_height = canvas_height / 6  # Максимум 6 недель в месяце

        # Рисуем заголовок с днями недели
        days = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]
        for i, day in enumerate(days):
            x = i * day_width
            self.calendar_canvas.create_rectangle(x, 0, x + day_width, day_height, fill="#444444", outline="black")
            self.calendar_canvas.create_text(x + day_width / 2, day_height / 2, text=day, fill="white",
                                             font=("Arial", 10))

        # Рисуем сетку календаря
        current_day = 1
        for week in range(6):
            for day in range(7):
                if (week == 0 and day < start_weekday) or current_day > total_days:
                    continue

                x = day * day_width
                y = (week + 1) * day_height

                # Рисуем ячейку дня
                self.calendar_canvas.create_rectangle(
                    x, y, x + day_width, y + day_height,
                    fill="#555555", outline="black"
                )

                # Подпись числа
                self.calendar_canvas.create_text(
                    x + 5, y + 5,
                    text=str(current_day),
                    anchor="nw", fill="white", font=("Arial", 10)
                )

                # Проверяем заказы на эту дату
                current_date = datetime(self.current_year, self.current_month, current_day).date()
                day_orders = [o for o in orders
                              if datetime.strptime(o[3], "%Y-%m-%d").date() == current_date]

                if day_orders:
                    # Разделяем заказы на выполненные и невыполненные
                    completed_orders = [o for o in day_orders if o[5] == "Завершен"]
                    active_orders = [o for o in day_orders if o[5] != "Завершен"]

                    # Рисуем индикаторы
                    if completed_orders:
                        # Зеленый для выполненных
                        self.calendar_canvas.create_oval(
                            x + day_width - 15, y + 5,
                            x + day_width - 5, y + 15,
                            fill="#4CAF50", outline="black"
                        )
                        self.calendar_canvas.create_text(
                            x + day_width - 10, y + 10,
                            text=str(len(completed_orders)),
                            fill="white", font=("Arial", 8)
                        )

                    if active_orders:
                        # Красный для активных
                        self.calendar_canvas.create_oval(
                            x + day_width - 35, y + 5,
                            x + day_width - 25, y + 15,
                            fill="#F44336", outline="black"
                        )
                        self.calendar_canvas.create_text(
                            x + day_width - 30, y + 10,
                            text=str(len(active_orders)),
                            fill="white", font=("Arial", 8)
                        )

                current_day += 1

    def on_calendar_click(self, event):
        """Обработчик клика по календарю"""
        canvas_width = self.calendar_canvas.winfo_width()
        canvas_height = self.calendar_canvas.winfo_height()
        day_width = canvas_width / 7
        day_height = canvas_height / 6

        # Определяем координаты клика
        day_col = int(event.x // day_width)
        day_row = int(event.y // day_height)

        # Если клик в заголовке или вне дней месяца
        if day_row == 0 or day_col < 0 or day_col > 6:
            return

        # Получаем первый день месяца
        first_day = datetime(self.current_year, self.current_month, 1).date()
        start_weekday = first_day.weekday()  # 0-пн, 6-вс

        # Вычисляем день месяца
        day_num = (day_row - 1) * 7 + day_col - start_weekday + 1
        last_day = (datetime(self.current_year, self.current_month + 1, 1) - timedelta(days=1)).day

        if day_num < 1 or day_num > last_day:
            return

        # Получаем заказы на эту дату
        current_date = datetime(self.current_year, self.current_month, day_num).date()
        orders = get_production_orders()
        day_orders = [o for o in orders
                      if datetime.strptime(o[3], "%Y-%m-%d").date() == current_date]

        if not day_orders:
            messagebox.showinfo("Информация", f"На {current_date.strftime('%d.%m.%Y')} нет заказов")
            return

        # Создаем окно с информацией о заказах
        self.show_day_orders(day_orders, current_date)

    def show_day_orders(self, orders, date):
        """Показывает заказы на выбранную дату"""
        dialog = tk.Toplevel(self)
        dialog.title(f"Заказы на {date.strftime('%d.%m.%Y')}")
        dialog.geometry("500x400")

        # Список заказов
        orders_frame = CTkFrame(dialog)
        orders_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        scrollbar = CTkScrollbar(orders_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        orders_listbox = tk.Listbox(
            orders_frame,
            bg="#333333",
            fg="white",
            font=("Arial", 12),
            yscrollcommand=scrollbar.set
        )
        orders_listbox.pack(fill=tk.BOTH, expand=True)
        scrollbar.configure(command=orders_listbox.yview)

        # Заполняем список
        for order in orders:
            order_id, name, customer, deadline, priority, status = order
            orders_listbox.insert(tk.END, f"{name} (Приоритет: {priority}, Статус: {status})")

        # Кнопка просмотра
        def view_order():
            selection = orders_listbox.curselection()
            if not selection:
                return

            selected_order = orders[selection[0]]
            self.show_order_details_by_id(selected_order[0])
            dialog.destroy()

        button_frame = CTkFrame(dialog)
        button_frame.pack(fill=tk.X, padx=10, pady=5)

        CTkButton(button_frame, text="Просмотреть", command=view_order).pack(side=tk.LEFT, padx=5)
        CTkButton(button_frame, text="Закрыть", command=dialog.destroy).pack(side=tk.RIGHT, padx=5)

    def show_order_details_by_id(self, order_id):
        """Показывает детали заказа по ID"""
        self.current_order_id = order_id
        orders = get_production_orders()

        for order in orders:
            if order[0] == order_id:
                order_id, name, customer, deadline, priority, status = order
                deadline_date = datetime.strptime(deadline, "%Y-%m-%d").strftime("%d.%m.%Y")

                details = f"Заказ №{order_id}\n\n"
                details += f"Название: {name}\n"
                details += f"Заказчик: {customer}\n"
                details += f"Срок выполнения: {deadline_date}\n"
                details += f"Приоритет: {priority}\n"
                details += f"Статус: {status}\n"

                self.order_details_text.config(state=tk.NORMAL)
                self.order_details_text.delete(1.0, tk.END)
                self.order_details_text.insert(tk.END, details)
                self.order_details_text.config(state=tk.DISABLED)

                # Загружаем стеклопакеты для этого заказа
                self.load_windows_for_order(order_id)
                break

    def show_prev_month(self):
        """Показывает предыдущий месяц"""
        self.current_month -= 1
        if self.current_month < 1:
            self.current_month = 12
            self.current_year -= 1
        self.update_calendar()

    def show_next_month(self):
        """Показывает следующий месяц"""
        self.current_month += 1
        if self.current_month > 12:
            self.current_month = 1
            self.current_year += 1
        self.update_calendar()

    def add_production_order(self):
        """Добавление нового производственного заказа"""
        name = self.order_name_entry.get()
        customer = self.customer_entry.get()
        deadline = self.deadline_entry.get()
        priority = self.priority_var.get()

        if not name or not deadline:
            messagebox.showerror("Ошибка", "Название и срок обязательны для заполнения")
            return

        try:
            deadline_date = datetime.strptime(deadline, "%d.%m.%y").date()
        except ValueError:
            try:
                deadline_date = datetime.strptime(deadline, "%d.%m.%Y").date()
            except ValueError:
                messagebox.showerror("Ошибка", "Неверный формат даты. Используйте ДД.ММ.ГГ или ДД.ММ.ГГГГ")
                return

        # Добавляем заказ в БД
        order_id = add_production_order(name, customer, deadline_date.strftime("%Y-%m-%d"), priority, "В ожидании")

        # Загружаем список заказов
        self.load_production_orders()

        # Устанавливаем текущий заказ
        self.current_order_id = order_id
        self.show_order_details_by_id(order_id)

        # Очищаем поля
        self.order_name_entry.delete(0, tk.END)
        self.customer_entry.delete(0, tk.END)
        self.deadline_entry.delete(0, tk.END)

        # Обновляем календарь
        self.update_calendar()

    def load_production_orders(self):
        """Загрузка списка производственных заказов"""
        orders = get_production_orders()
        self.orders_listbox.delete(0, tk.END)

        for order in orders:
            order_id, name, customer, deadline, priority, status = order
            deadline_date = datetime.strptime(deadline, "%Y-%m-%d").strftime("%d.%m.%Y")
            self.orders_listbox.insert(tk.END, f"{order_id}: {name} ({status})")

    def show_order_details(self, event=None):
        """Отображение деталей выбранного заказа"""
        selection = self.orders_listbox.curselection()
        if not selection:
            return

        # Получаем ID заказа из строки (формат "ID: Название")
        order_text = self.orders_listbox.get(selection[0])
        order_id = order_text.split(":")[0].strip()

        self.current_order_id = order_id
        orders = get_production_orders()

        for order in orders:
            if str(order[0]) == order_id:
                order_id, name, customer, deadline, priority, status = order
                deadline_date = datetime.strptime(deadline, "%Y-%m-%d").strftime("%d.%m.%Y")

                details = f"Заказ №{order_id}\n\n"
                details += f"Название: {name}\n"
                details += f"Заказчик: {customer}\n"
                details += f"Срок выполнения: {deadline_date}\n"
                details += f"Приоритет: {priority}\n"
                details += f"Статус: {status}\n"

                # Обновляем текстовое поле
                self.order_details_text.config(state=tk.NORMAL)
                self.order_details_text.delete(1.0, tk.END)
                self.order_details_text.insert(tk.END, details)
                self.order_details_text.config(state=tk.DISABLED)

                # Загружаем связанные стеклопакеты
                self.load_windows_for_order(order_id)
                break

    def load_windows_for_order(self, order_id):
        """Загрузка списка стеклопакетов для заказа"""
        # Очищаем таблицу
        for item in self.windows_tree.get_children():
            self.windows_tree.delete(item)

        # Загружаем данные из БД
        windows = get_windows_for_production_order(order_id)
        for window in windows:
            self.windows_tree.insert("", tk.END, values=window)

    def add_window_to_order(self):
        """Добавление стеклопакета к заказу"""
        if not self.current_order_id:
            messagebox.showwarning("Предупреждение", "Сначала выберите заказ")
            return

        # Создаем диалоговое окно
        self.window_dialog = tk.Toplevel(self)
        self.window_dialog.title("Добавить стеклопакет")
        self.window_dialog.geometry("300x250")
        self.window_dialog.protocol("WM_DELETE_WINDOW", self.close_window_dialog)

        # Переменные для хранения значений
        self.width_var = tk.StringVar()
        self.height_var = tk.StringVar()
        self.quantity_var = tk.StringVar()
        self.type_var = tk.StringVar(value="Однокамерный")

        # Элементы управления
        CTkLabel(self.window_dialog, text="Ширина (мм):").pack(pady=5)
        width_entry = CTkEntry(self.window_dialog, textvariable=self.width_var)
        width_entry.pack(pady=5)
        width_entry.focus_set()

        CTkLabel(self.window_dialog, text="Высота (мм):").pack(pady=5)
        height_entry = CTkEntry(self.window_dialog, textvariable=self.height_var)
        height_entry.pack(pady=5)

        CTkLabel(self.window_dialog, text="Количество:").pack(pady=5)
        quantity_entry = CTkEntry(self.window_dialog, textvariable=self.quantity_var)
        quantity_entry.pack(pady=5)

        CTkLabel(self.window_dialog, text="Тип:").pack(pady=5)
        type_combo = CTkComboBox(self.window_dialog,
                                 values=["Однокамерный", "Двухкамерный"],
                                 variable=self.type_var)
        type_combo.pack(pady=5)

        # Кнопки
        button_frame = CTkFrame(self.window_dialog)
        button_frame.pack(pady=10)

        CTkButton(button_frame, text="Сохранить и добавить еще",
                  command=self.save_window_and_continue).pack(side=tk.LEFT, padx=5)
        CTkButton(button_frame, text="Сохранить и закрыть",
                  command=self.save_window_and_close).pack(side=tk.LEFT, padx=5)
        CTkButton(button_frame, text="Отмена",
                  command=self.close_window_dialog).pack(side=tk.LEFT, padx=5)

    def save_window_and_continue(self):
        """Сохраняет текущий стеклопакет и очищает поля для ввода следующего"""
        try:
            width = int(self.width_var.get())
            height = int(self.height_var.get())
            quantity = int(self.quantity_var.get())
            window_type = self.type_var.get()

            if width <= 0 or height <= 0:
                messagebox.showerror("Ошибка", "Размеры должны быть положительными числами")
                return

            if quantity <= 0:
                messagebox.showerror("Ошибка", "Количество должно быть положительным")
                return

            add_window_to_production_order(self.current_order_id, window_type, width, height, quantity)
            self.load_windows_for_order(self.current_order_id)

            # Очищаем поля для ввода следующего значения
            self.width_var.set("")
            self.height_var.set("")
            self.quantity_var.set("")
            self.window_dialog.focus_set()  # Возвращаем фокус в окно

        except ValueError:
            messagebox.showerror("Ошибка", "Введите корректные числовые значения")

    def save_window_and_close(self):
        """Сохраняет стеклопакет и закрывает окно"""
        self.save_window_and_continue()
        self.close_window_dialog()

    def close_window_dialog(self):
        """Закрывает окно добавления стеклопакетов"""
        if hasattr(self, 'window_dialog') and self.window_dialog:
            self.window_dialog.destroy()

    def delete_window_from_order(self):
        """Удаление стеклопакета из заказа"""
        if not self.current_order_id:
            messagebox.showwarning("Предупреждение", "Сначала выберите заказ")
            return

        selection = self.windows_tree.selection()
        if not selection:
            messagebox.showwarning("Предупреждение", "Выберите стеклопакет для удаления")
            return

        if messagebox.askyesno("Подтверждение", "Удалить выбранный стеклопакет?"):
            window_id = self.windows_tree.item(selection[0], "values")[0]
            delete_window_from_production_order(window_id)
            self.load_windows_for_order(self.current_order_id)

    def change_order_status(self, new_status):
        """Изменение статуса заказа"""
        if not self.current_order_id:
            messagebox.showwarning("Предупреждение", "Выберите заказ")
            return

        update_production_order_status(self.current_order_id, new_status)
        self.load_production_orders()
        self.show_order_details(None)
        self.update_calendar()

    def delete_order(self):
        """Удаление производственного заказа"""
        if not self.current_order_id:
            messagebox.showwarning("Предупреждение", "Выберите заказ")
            return

        if messagebox.askyesno("Подтверждение", "Удалить выбранный заказ?"):
            delete_production_order(self.current_order_id)
            self.load_production_orders()
            self.order_details_text.config(state=tk.NORMAL)
            self.order_details_text.delete(1.0, tk.END)
            self.order_details_text.config(state=tk.DISABLED)
            self.current_order_id = None

            # Очищаем таблицу стеклопакетов
            for item in self.windows_tree.get_children():
                self.windows_tree.delete(item)
        self.update_calendar()

    def import_order_from_excel(self):
        """Импорт заказа из Excel файла"""
        file_path = filedialog.askopenfilename(
            title="Выберите файл заказа",
            filetypes=[("Excel files", "*.xls *.xlsx"), ("All files", "*.*")]
        )

        if not file_path:
            return

        try:
            # Парсим данные из Excel
            order_data = self.parse_excel_order(file_path)

            # Заполняем поля формы
            self.order_name_entry.delete(0, tk.END)
            self.order_name_entry.insert(0, order_data['order_name'])

            self.customer_entry.delete(0, tk.END)
            self.customer_entry.insert(0, order_data['customer'])

            self.deadline_entry.delete(0, tk.END)
            self.deadline_entry.insert(0, order_data['deadline'])

            # Устанавливаем приоритет (можно добавить логику определения из файла)
            self.priority_var.set("Средний")

            # Показываем сообщение об успехе
            messagebox.showinfo("Успех", "Данные заказа успешно загружены из файла")

        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось загрузить данные из файла:\n{str(e)}")

    def parse_excel_order(self, file_path):
        """Парсинг данных заказа из Excel файла (поддержка .xls и .xlsx)"""
        try:
            # Определяем расширение файла
            ext = file_path.split('.')[-1].lower()

            order_data = {
                'order_name': '',
                'customer': '',
                'deadline': '',
                'windows': []
            }

            if ext == 'xls':
                # Используем xlrd для старых .xls файлов
                import xlrd
                book = xlrd.open_workbook(file_path)
                sheet = book.sheet_by_index(0)

                # Парсим основные данные заказа
                for row_idx in range(sheet.nrows):
                    row = sheet.row_values(row_idx)
                    # print("row ", row)

                    # Номер заказа (ищем строку, содержащую "Заказ №")
                    if not order_data['order_name'] and row and len(row) > 1:
                        cell_value = str(row[1]).strip()
                        if "Заказ №" in cell_value:
                            order_data['order_name'] = cell_value
                            print(cell_value)

                    # Заказчик (ищем строку, где первый столбец содержит "Заказчик:" или "Çàêàç÷èê:")
                    if not order_data['customer'] and row and len(row) > 1:
                        cell_value = str(row[1]).strip()
                        if "Заказчик:" in cell_value:
                            customer = str(row[8]).strip() if len(row) > 8 and row[8] else ""
                            # Очищаем от лишней информации (телефонов и т.д.)
                            customer = customer.split(',')[0].split('тел.:')[0].strip()
                            order_data['customer'] = customer
                            print(customer)

                    # Дата доставки (ищем строку, где первый столбец содержит "Дата доставки:" или "Äàòà äîñòàâêè:")
                    if not order_data['deadline'] and row and len(row) > 1:
                        cell_value = str(row[17]).strip()
                        if "Дата изготовления:" in cell_value:
                            if len(row) > 17 and row[24]:
                                date_str = str(row[24]).split('\\')[0].strip()
                                # Пропускаем некорректные даты (например, ". .")
                                if date_str and date_str != '. .':
                                    # Преобразуем дату из формата "09.04.25" в "09.04.2025"
                                    if len(date_str.split('.')) == 3 and len(date_str.split('.')[2]) == 2:
                                        day, month, year = date_str.split('.')
                                        date_str = f"{day}.{month}.20{year}"
                                    order_data['deadline'] = date_str
                                    print(date_str)

                # Парсим список стеклопакетов (начинаем с строки, где есть заголовок "№")
                start_parsing = False
                one_skip = False
                for row_idx in range(sheet.nrows):
                    if one_skip:
                        one_skip = False
                        continue
                    row = sheet.row_values(row_idx)
                    # print("row steklo ", row)
                    # Ищем начало таблицы со стеклопакетами
                    if row and len(row) > 1:
                        first_cell = str(row[1]).strip()
                        if first_cell in ["№"] and "Поз" in str(row[3]):
                            start_parsing = True
                            one_skip = True
                            continue

                    if start_parsing and row and len(row) > 14:
                        try:
                            # Проверяем, что первая ячейка содержит число (номер строки)
                            if row[1] and row[5]:
                                # Обрабатываем размеры (удаляем пробелы и разделяем на ширину/высоту)
                                size_str = str(row[15]).replace(' ', '') if len(row) > 15 and row[15] else "0x0"
                                size_parts = size_str.split('x')

                                # Получаем количество (может быть в разных столбцах)
                                quantity = 1
                                if len(row) > 20 and row[20]:
                                    try:
                                        quantity = int(float(row[20]))
                                    except:
                                        quantity = 1

                                window_data = {
                                    'type': str(row[5]).strip(),
                                    'width': int(float(size_parts[0])) if size_parts[0] else 0,
                                    'height': int(float(size_parts[1])) if len(size_parts) > 1 and size_parts[1] else 0,
                                    'quantity': quantity
                                }
                                order_data['windows'].append(window_data)
                                # print(window_data, "     dsfsd")
                                # for key, value in window_data.items():
                                #     print(f"{key}: {value}")
                            else:
                                start_parsing = False
                        except (ValueError, IndexError, AttributeError) as e:
                            print(f"Ошибка при парсинге строки {row_idx}: {e}")
                            continue
                print(order_data['windows'])


            elif ext == 'xlsx':
                # Используем openpyxl для новых .xlsx файлов
                import openpyxl
                wb = openpyxl.load_workbook(file_path)
                sheet = wb.active

                # Парсим основные данные заказа
                for row in sheet.iter_rows(values_only=True):
                    # Номер заказа
                    if not order_data['order_name'] and row and len(row) > 1:
                        cell_value = str(row[1]).strip()
                        if "Заказ №" in cell_value:
                            order_data['order_name'] = cell_value

                    # Заказчик
                    if not order_data['customer'] and row and len(row) > 1:
                        cell_value = str(row[1]).strip()
                        if "Заказчик:" in cell_value:
                            customer = str(row[7]).strip() if len(row) > 7 and row[7] else ""
                            customer = customer.split(',')[0].split('тел.:')[0].strip()
                            order_data['customer'] = customer

                    # Дата доставки
                    if not order_data['deadline'] and row and len(row) > 1:
                        cell_value = str(row[1]).strip()
                        if "Дата доставки:" in cell_value:
                            if len(row) > 7 and row[7]:
                                date_str = str(row[7]).split('\\')[0].strip()
                                if date_str and date_str != '. .':
                                    if len(date_str.split('.')) == 3 and len(date_str.split('.')[2]) == 2:
                                        day, month, year = date_str.split('.')
                                        date_str = f"{day}.{month}.20{year}"
                                    order_data['deadline'] = date_str

                # Парсим список стеклопакетов
                start_parsing = False
                for row in sheet.iter_rows(values_only=True):
                    # Ищем начало таблицы со стеклопакетами
                    if row and len(row) > 1:
                        first_cell = str(row[1]).strip()
                        if first_cell in ["№", "¹"] and "Поз" in str(row[3]):
                            start_parsing = True
                            continue

                    if start_parsing and row and len(row) > 14:
                        try:
                            if isinstance(row[0], (int, float)) and row[4] and any(
                                    x in str(row[4]) for x in ["СПД", "СПО", "ÑÏÄ", "ÑÏÎ"]):
                                size_str = str(row[14]).replace(' ', '') if len(row) > 14 and row[14] else "0x0"
                                size_parts = size_str.split('x')

                                quantity = 1
                                if len(row) > 19 and row[19]:
                                    try:
                                        quantity = int(float(row[19]))
                                    except:
                                        quantity = 1

                                window_data = {
                                    'type': str(row[4]).strip(),
                                    'width': int(float(size_parts[0])) if size_parts[0] else 0,
                                    'height': int(float(size_parts[1])) if len(size_parts) > 1 and size_parts[1] else 0,
                                    'quantity': quantity
                                }
                                order_data['windows'].append(window_data)
                        except (ValueError, IndexError, AttributeError) as e:
                            print(f"Ошибка при парсинге строки: {e}")
                            continue
            else:
                raise Exception("Неподдерживаемый формат файла. Используйте .xls или .xlsx")

            return order_data

        except Exception as e:
            raise Exception(f"Ошибка при чтении файла: {str(e)}")


class CuttingOptimizer(CTk):
    def __init__(self):
        super().__init__()
        self.title("Оптимизация раскроя заготовки")
        self.geometry("1150x800")
        self.withdraw()

        # Создаем вкладки
        self.tab_control = ttk.Notebook(self)

        # Вкладка для раскроя рамки
        self.frame_tab = FrameCuttingTab(self)
        self.tab_control.add(self.frame_tab, text="Раскрой рамки")

        # Вкладка для раскроя стекла
        self.glass_tab = GlassCuttingTab(self)
        self.tab_control.add(self.glass_tab, text="Раскрой стекла")

        # Вкладка для планирования производства
        self.planning_tab = ProductionPlanningTab(self)
        self.tab_control.add(self.planning_tab, text="Планирование производства")

        self.tab_control.pack(expand=1, fill="both")


        # Привязываем обработчики событий
        self.frame_tab.card_listbox.bind('<<ListboxSelect>>', self.frame_tab.display_card_details)
        self.glass_tab.card_listbox.bind('<<ListboxSelect>>', self.glass_tab.display_card_details)

    def on_orders_updated(self):
        """Обновляем списки заказов в обеих вкладках"""
        self.frame_tab.load_orders_from_db()
        self.glass_tab.load_orders_from_db()


if __name__ == "__main__":
    create_database()
    app = CuttingOptimizer()
    auth_window = AuthWindow(app)

    auth_window.mainloop()