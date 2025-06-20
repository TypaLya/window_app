import tkinter as tk
from tkinter import ttk, filedialog
from datetime import datetime, timedelta

import openpyxl
import xlrd
from customtkinter import CTkLabel, CTkEntry, CTkButton, CTkFrame, CTkScrollbar, CTkComboBox, CTkScrollableFrame
from tkinter import messagebox
from database import (add_production_order, get_production_orders,
                      update_production_order_status, delete_production_order,
                      add_window_to_production_order, get_windows_for_production_order,
                      delete_window_from_production_order, delete_material_from_production_order,
                      add_material_to_production_order, get_materials_for_production_order, get_all_cutting_materials,
                      get_all_main_glass_materials, get_all_triplex_materials, get_all_window_materials,
                      get_all_film_materials, get_all_component_materials)


def parse_excel_order(file_path):
    """Парсинг данных заказа из Excel файла (поддержка .xls и .xlsx)"""
    try:
        # Определяем расширение файла
        ext = file_path.split('.')[-1].lower()

        order_data = {
            'order_name': '',
            'customer': '',
            'deadline': '',
            'windows': [],
            'materials': []
        }

        if ext == 'xls':
            book = xlrd.open_workbook(file_path)
            sheet = book.sheet_by_index(0)

            # Парсим основные данные заказа
            for row_idx in range(sheet.nrows):
                row = sheet.row_values(row_idx)
                # print("row ", row)

                # Номер заказа
                if not order_data['order_name'] and row and len(row) > 1:
                    cell_value = str(row[1]).strip()
                    if "Заказ №" in cell_value:
                        order_data['order_name'] = cell_value
                        # print(cell_value)

                # Заказчик
                if not order_data['customer'] and row and len(row) > 1:
                    cell_value = str(row[1]).strip()
                    if "Заказчик:" in cell_value:
                        customer = str(row[8]).strip() if len(row) > 8 and row[8] else ""
                        customer = customer.split(',')[0].split('тел.:')[0].strip()
                        order_data['customer'] = customer
                        # print(customer)

                # Дата доставки
                if not order_data['deadline'] and row and len(row) > 1:
                    cell_value = str(row[17]).strip()
                    if "Дата изготовления:" in cell_value:
                        if len(row) > 17 and row[24]:
                            date_str = str(row[24]).split('\\')[0].strip()
                            if date_str and date_str != '. .':
                                if len(date_str.split('.')) == 3 and len(date_str.split('.')[2]) == 2:
                                    day, month, year = date_str.split('.')
                                    date_str = f"{day}.{month}.20{year}"
                                order_data['deadline'] = date_str
                                # print(date_str)

            # Парсим список стеклопакетов
            start_parsing = False
            start_parsing_materials = False
            one_skip = False
            for row_idx in range(sheet.nrows):
                if one_skip:
                    one_skip = False
                    continue
                row = sheet.row_values(row_idx)
                # print("row steklo ", row)
                # Ищем начало таблицы со стеклопакетами
                if row and len(row) > 1:
                    first_cell = str(row[1]).strip()
                    if first_cell in ["№"] and "Поз" in str(row[3]):
                        start_parsing = True
                        one_skip = True
                        continue
                    if "Расход комплектующих" in first_cell:
                        start_parsing_materials = True
                        continue

                if start_parsing and row and len(row) > 14:
                    try:
                        if row[1] and row[5]:
                            size_str = str(row[15]).replace(' ', '') if len(row) > 15 and row[15] else "0x0"
                            size_parts = size_str.split('x')

                            quantity = 1
                            if len(row) > 20 and row[20]:
                                try:
                                    quantity = int(float(row[20]))
                                except:
                                    quantity = 1

                            window_data = {
                                'type': str(row[5]).strip(),
                                'width': int(float(size_parts[0])) if size_parts[0] else 0,
                                'height': int(float(size_parts[1])) if len(size_parts) > 1 and size_parts[1] else 0,
                                'quantity': quantity
                            }
                            order_data['windows'].append(window_data)
                            # print(window_data, "     dsfsd")
                            # for key, value in window_data.items():
                            #     print(f"{key}: {value}")
                        else:
                            start_parsing = False
                    except (ValueError, IndexError, AttributeError) as e:
                        print(f"Ошибка при парсинге строки {row_idx}: {e}")
                        continue
                if start_parsing_materials and row and len(row) > 30:
                    try:
                        if row[1] and row[25] and row[30]:
                            material_str = str(row[1]).strip()
                            material_amount = float(row[25])
                            materiaL_dimension = str(row[30]).strip()

                            material_data = {
                                'type': material_str,
                                'amount': material_amount,
                                'dimension': materiaL_dimension
                            }

                            order_data['materials'].append(material_data)
                            # print(material_data, "     dsfsd")
                            # for key, value in material_data.items():
                                # print(f"{key}: {value}")
                        else:
                            start_parsing_materials = False
                    except (ValueError, IndexError, AttributeError) as e:
                        print(f"Ошибка при парсинге строки {row_idx}: {e}")
                        continue


        elif ext == 'xlsx':
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.active

            # Парсим основные данные заказа
            for row in sheet.iter_rows(values_only=True):
                # Номер заказа
                if not order_data['order_name'] and row and len(row) > 1:
                    cell_value = str(row[1]).strip()
                    if "Заказ №" in cell_value:
                        order_data['order_name'] = cell_value

                # Заказчик
                if not order_data['customer'] and row and len(row) > 1:
                    cell_value = str(row[1]).strip()
                    if "Заказчик:" in cell_value:
                        customer = str(row[7]).strip() if len(row) > 7 and row[7] else ""
                        customer = customer.split(',')[0].split('тел.:')[0].strip()
                        order_data['customer'] = customer

                # Дата доставки
                if not order_data['deadline'] and row and len(row) > 1:
                    cell_value = str(row[1]).strip()
                    if "Дата доставки:" in cell_value:
                        if len(row) > 7 and row[7]:
                            date_str = str(row[7]).split('\\')[0].strip()
                            if date_str and date_str != '. .':
                                if len(date_str.split('.')) == 3 and len(date_str.split('.')[2]) == 2:
                                    day, month, year = date_str.split('.')
                                    date_str = f"{day}.{month}.20{year}"
                                order_data['deadline'] = date_str

            # Парсим список стеклопакетов
            start_parsing = False
            for row in sheet.iter_rows(values_only=True):
                # Ищем начало таблицы со стеклопакетами
                if row and len(row) > 1:
                    first_cell = str(row[1]).strip()
                    if first_cell in ["№", "¹"] and "Поз" in str(row[3]):
                        start_parsing = True
                        continue

                if start_parsing and row and len(row) > 14:
                    try:
                        if isinstance(row[0], (int, float)) and row[4] and any(
                                x in str(row[4]) for x in ["СПД", "СПО", "ÑÏÄ", "ÑÏÎ"]):
                            size_str = str(row[14]).replace(' ', '') if len(row) > 14 and row[14] else "0x0"
                            size_parts = size_str.split('x')

                            quantity = 1
                            if len(row) > 19 and row[19]:
                                try:
                                    quantity = int(float(row[19]))
                                except:
                                    quantity = 1

                            window_data = {
                                'type': str(row[4]).strip(),
                                'width': int(float(size_parts[0])) if size_parts[0] else 0,
                                'height': int(float(size_parts[1])) if len(size_parts) > 1 and size_parts[1] else 0,
                                'quantity': quantity
                            }
                            order_data['windows'].append(window_data)
                    except (ValueError, IndexError, AttributeError) as e:
                        print(f"Ошибка при парсинге строки: {e}")
                        continue
        else:
            raise Exception("Неподдерживаемый формат файла. Используйте .xls или .xlsx")

        return order_data

    except Exception as e:
        raise Exception(f"Ошибка при чтении файла: {str(e)}")


class ProductionPlanningTab(CTkFrame):
    def __init__(self, parent):
        super().__init__(parent)
        self.parent = parent
        self.current_order_id = None
        self._first_draw = True  # Флаг первого отображения

        # Основные фреймы
        self.left_frame = CTkFrame(self, width=300)
        self.left_frame.pack(side=tk.LEFT, fill=tk.Y, padx=10, pady=10)

        self.right_frame = CTkFrame(self)
        self.right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Форма добавления заказа
        self.add_order_frame = CTkFrame(self.left_frame)
        self.add_order_frame.pack(fill=tk.X, pady=5)

        CTkLabel(self.add_order_frame, text="Новый производственный заказ").pack(pady=5)

        # Поля ввода
        self.order_name_entry = CTkEntry(self.add_order_frame, placeholder_text="Название заказа")
        self.order_name_entry.pack(fill=tk.X, pady=5)

        self.customer_entry = CTkEntry(self.add_order_frame, placeholder_text="Заказчик")
        self.customer_entry.pack(fill=tk.X, pady=5)

        self.deadline_entry = CTkEntry(self.add_order_frame, placeholder_text="Срок (ДД.ММ.ГГГГ)")
        self.deadline_entry.pack(fill=tk.X, pady=5)

        self.priority_var = tk.StringVar(value="Средний")
        self.priority_combo = CTkComboBox(self.add_order_frame,
                                          values=["Низкий", "Средний", "Высокий", "Критичный"],
                                          variable=self.priority_var)
        self.priority_combo.pack(fill=tk.X, pady=5)

        self.add_button = CTkButton(self.add_order_frame, text="Добавить заказ", command=self.add_production_order)
        self.add_button.pack(fill=tk.X, pady=5)

        self.import_button = CTkButton(self.add_order_frame, text="Импорт заказа",
                                       command=self.import_order_from_excel)
        self.import_button.pack(fill=tk.X, pady=5)

        # Список заказов
        self.orders_frame = CTkFrame(self.left_frame)
        self.orders_frame.pack(fill=tk.BOTH, expand=True, pady=5)

        self.orders_listbox = tk.Listbox(self.orders_frame, bg="#333333", fg="white", font=("Arial", 12))
        self.orders_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.orders_listbox.bind('<<ListboxSelect>>', self.show_order_details)

        self.scrollbar = CTkScrollbar(self.orders_frame)
        self.scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.orders_listbox.config(yscrollcommand=self.scrollbar.set)
        self.scrollbar.configure(command=self.orders_listbox.yview)
        self.load_production_orders()

        # Кнопки управления
        self.control_frame = CTkFrame(self.left_frame)
        self.control_frame.pack(fill=tk.X, pady=5)

        self.start_button = CTkButton(self.control_frame, text="В работу",
                                      command=lambda: self.change_order_status("В работе"))
        self.start_button.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=2)

        self.complete_button = CTkButton(self.control_frame, text="Завершить",
                                         command=lambda: self.change_order_status("Завершен"))
        self.complete_button.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=2)

        self.delete_button = CTkButton(self.control_frame, text="Удалить", command=self.delete_order)
        self.delete_button.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=2)

        # Детали заказа
        self.details_frame = CTkFrame(self.right_frame)
        self.details_frame.pack(fill=tk.BOTH, expand=True, pady=5)

        # Вкладки для деталей заказа
        self.details_notebook = ttk.Notebook(self.details_frame)
        self.details_notebook.pack(fill=tk.BOTH, expand=True)

        self.info_tab = CTkFrame(self.details_notebook)
        self.details_notebook.add(self.info_tab, text="Информация")

        self.order_details_text = tk.Text(self.info_tab, bg="#333333", fg="white", font=("Arial", 12))
        self.order_details_text.pack(fill=tk.BOTH, expand=True)

        # Вкладка со стеклопакетами
        self.windows_tab = CTkFrame(self.details_notebook)
        self.details_notebook.add(self.windows_tab, text="Стеклопакеты")

        # Кнопки управления стеклопакетами
        self.windows_control_frame = CTkFrame(self.windows_tab)
        self.windows_control_frame.pack(fill=tk.X, pady=5)

        self.add_window_button = CTkButton(self.windows_control_frame, text="Добавить стеклопакет",
                                           command=self.add_window_to_order)
        self.add_window_button.pack(side=tk.LEFT, padx=5)

        self.delete_window_button = CTkButton(self.windows_control_frame, text="Удалить стеклопакет",
                                              command=self.delete_window_from_order)
        self.delete_window_button.pack(side=tk.LEFT, padx=5)

        # Таблица стеклопакетов
        self.windows_tree = ttk.Treeview(self.windows_tab, columns=("id", "type", "width", "height", "quantity"),
                                         show="headings")
        self.windows_tree.heading("id", text="ID")
        self.windows_tree.heading("type", text="Тип")
        self.windows_tree.heading("width", text="Ширина (мм)")
        self.windows_tree.heading("height", text="Высота (мм)")
        self.windows_tree.heading("quantity", text="Количество")
        self.windows_tree.column("id", width=50)
        self.windows_tree.column("type", width=150)
        self.windows_tree.column("width", width=75)
        self.windows_tree.column("height", width=75)
        self.windows_tree.column("quantity", width=50)

        scrollbar = ttk.Scrollbar(self.windows_tab, orient="vertical", command=self.windows_tree.yview)
        self.windows_tree.configure(yscrollcommand=scrollbar.set)

        self.windows_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # Вкладка с расходами
        self.materials_tab = CTkFrame(self.details_notebook)
        self.details_notebook.add(self.materials_tab, text="Расходы")

        # Кнопки управления материалами
        self.materials_control_frame = CTkFrame(self.materials_tab)
        self.materials_control_frame.pack(fill=tk.X, pady=5)

        self.add_material_button = CTkButton(self.materials_control_frame,
                                             text="Добавить материал",
                                             command=self.add_material_to_order)
        self.add_material_button.pack(side=tk.LEFT, padx=5)

        self.delete_material_button = CTkButton(self.materials_control_frame,
                                                text="Удалить материал",
                                                command=self.delete_material_from_order)
        self.delete_material_button.pack(side=tk.LEFT, padx=5)

        # Таблица материалов
        self.materials_tree = ttk.Treeview(self.materials_tab,
                                           columns=("num", "type", "amount", "dimension"),
                                           show="headings")
        self.materials_tree.heading("num", text="№")
        self.materials_tree.heading("type", text="Материал")
        self.materials_tree.heading("amount", text="Количество")
        self.materials_tree.heading("dimension", text="Ед. изм.")
        self.materials_tree.column("num", width=50, anchor='center')
        self.materials_tree.column("type", width=200)
        self.materials_tree.column("amount", width=100, anchor='center')
        self.materials_tree.column("dimension", width=100, anchor='center')

        materials_scrollbar = ttk.Scrollbar(self.materials_tab,
                                            orient="vertical",
                                            command=self.materials_tree.yview)
        self.materials_tree.configure(yscrollcommand=materials_scrollbar.set)

        self.materials_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        materials_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # Вкладка со складом
        self.warehouse_tab = CTkFrame(self.details_notebook)
        self.details_notebook.add(self.warehouse_tab, text="Склад")

        # Кнопки управления складом
        self.warehouse_control_frame = CTkFrame(self.warehouse_tab)
        self.warehouse_control_frame.pack(fill=tk.X, pady=5)

        self.refresh_button = CTkButton(self.warehouse_control_frame,
                                        text="Обновить",
                                        command=self.load_warehouse_data)
        self.refresh_button.pack(side=tk.LEFT, padx=5)

        # Таблица склада
        self.warehouse_tree = ttk.Treeview(self.warehouse_tab,
                                           columns=("type", "amount", "dimension"),
                                           show="headings")
        self.warehouse_tree.heading("type", text="Материал")
        self.warehouse_tree.heading("amount", text="Используется")
        self.warehouse_tree.heading("dimension", text="На складе")

        self.warehouse_tree.column("type", width=250)
        self.warehouse_tree.column("amount", width=150, anchor='center')
        self.warehouse_tree.column("dimension", width=100, anchor='center')

        scrollbar = ttk.Scrollbar(self.warehouse_tab,
                                  orient="vertical",
                                  command=self.warehouse_tree.yview)
        self.warehouse_tree.configure(yscrollcommand=scrollbar.set)

        self.warehouse_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # Загружаем данные склада при инициализации
        self.load_warehouse_data()


        # Календарь производства
        self.calendar_frame = CTkFrame(self.right_frame)
        self.calendar_frame.pack(fill=tk.BOTH, expand=True, pady=5)

       # Загрузка данных
        self.load_production_orders()

        # Загрузка данных и отрисовка календаря
        self.after(100, self.update_calendar)  # Запускаем через небольшой таймаут
        self.bind("<Visibility>", self.on_visibility_changed)
        self.bind("<Configure>", self.on_resize)  # Обработчик изменения размеров

        # Панель навигации по месяцам
        self.calendar_nav_frame = CTkFrame(self.calendar_frame)
        self.calendar_nav_frame.pack(fill=tk.X, pady=5)

        self.prev_month_btn = CTkButton(self.calendar_nav_frame, text="<", width=30,
                                        command=self.show_prev_month)
        self.prev_month_btn.pack(side=tk.LEFT, padx=5)

        self.month_label = CTkLabel(self.calendar_nav_frame, text="", font=("Arial", 12))
        self.month_label.pack(side=tk.LEFT, expand=True)

        self.next_month_btn = CTkButton(self.calendar_nav_frame, text=">", width=30,
                                        command=self.show_next_month)
        self.next_month_btn.pack(side=tk.RIGHT, padx=5)

        # Холст календаря
        self.calendar_canvas = tk.Canvas(self.calendar_frame, bg="white")
        self.calendar_canvas.pack(fill=tk.BOTH, expand=True)
        self.calendar_canvas.bind("<Button-1>", self.on_calendar_click)

        # Текущий месяц и год для отображения
        self.current_date = datetime.now().date()
        self.current_month = self.current_date.month
        self.current_year = self.current_date.year

        # Загрузка данных и отрисовка календаря
        self.update_calendar()

    def on_visibility_changed(self, event):
        """Обработчик изменения видимости вкладки"""
        if self._first_draw and self.winfo_ismapped():
            self._first_draw = False
            self.after(100, self.update_calendar)

    def on_resize(self, event):
        """Обработчик изменения размеров окна"""
        if self.winfo_ismapped():
            self.after(100, self.update_calendar)

    def update_calendar(self):
        """Адаптивный календарь с правильным масштабированием"""
        if not self.winfo_ismapped() or not self.calendar_canvas.winfo_exists():
            return

        # Очищаем и настраиваем холст
        self.calendar_canvas.delete("all")
        month_names = ["", "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
                       "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"]
        self.month_label.configure(text=f"{month_names[self.current_month]} {self.current_year}")

        # Получаем актуальные размеры
        canvas_width = self.calendar_canvas.winfo_width()
        canvas_height = self.calendar_canvas.winfo_height()

        # Минимальные размеры
        if canvas_width < 300 or canvas_height < 200:
            canvas_width = 700
            canvas_height = 500
            self.calendar_canvas.config(width=canvas_width, height=canvas_height)

        # Рассчитываем дни месяца
        first_day = datetime(self.current_year, self.current_month, 1).date()
        if self.current_month == 12:
            last_day = datetime(self.current_year + 1, 1, 1).date() - timedelta(days=1)
        else:
            last_day = datetime(self.current_year, self.current_month + 1, 1).date() - timedelta(days=1)

        total_days = last_day.day
        start_weekday = first_day.weekday()
        today = datetime.now().date()

        # Рассчитываем необходимое количество строк (5 или 6 недель)
        num_weeks = 6 if (start_weekday + total_days) > 35 else 5

        # Размеры ячеек с отступами
        header_height = 30
        day_height = (canvas_height - header_height) / num_weeks
        day_width = canvas_width / 7

        # Рисуем заголовок дней недели
        days = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]
        for i, day in enumerate(days):
            x = i * day_width
            self.calendar_canvas.create_rectangle(
                x, 0, x + day_width, header_height,
                fill="#444444", outline="black"
            )
            self.calendar_canvas.create_text(
                x + day_width / 2, header_height / 2,
                text=day, fill="white", font=("Arial", 10, "bold")
            )

        # Получаем все заказы один раз
        all_orders = get_production_orders()

        # Рисуем дни месяца
        current_day = 1
        for week in range(num_weeks):
            for day in range(7):
                if (week == 0 and day < start_weekday) or current_day > total_days:
                    continue

                x = day * day_width
                y = header_height + week * day_height
                current_date = datetime(self.current_year, self.current_month, current_day).date()

                # Ячейка дня
                self.calendar_canvas.create_rectangle(
                    x, y, x + day_width, y + day_height,
                    fill="#8ab4f8" if current_date == today else "#555555",
                    outline="black"
                )

                # Число дня
                self.calendar_canvas.create_text(
                    x + 5, y + 5,
                    text=str(current_day),
                    anchor="nw", fill="white", font=("Arial", 10)
                )

                # Фильтруем заказы для текущей даты
                day_orders = [o for o in all_orders
                              if datetime.strptime(o[3], "%Y-%m-%d").date() == current_date]

                if day_orders:
                    active = sum(1 for o in day_orders if o[5] != "Завершен")
                    completed = sum(1 for o in day_orders if o[5] == "Завершен")

                    if active:
                        self.calendar_canvas.create_oval(
                            x + day_width - 35, y + 5,
                            x + day_width - 25, y + 15,
                            fill="#F44336", outline="black"
                        )
                        self.calendar_canvas.create_text(
                            x + day_width - 30, y + 10,
                            text=str(active), fill="white", font=("Arial", 8)
                        )

                    if completed:
                        self.calendar_canvas.create_oval(
                            x + day_width - 15, y + 5,
                            x + day_width - 5, y + 15,
                            fill="#4CAF50", outline="black"
                        )
                        self.calendar_canvas.create_text(
                            x + day_width - 10, y + 10,
                            text=str(completed), fill="white", font=("Arial", 8)
                        )

                current_day += 1

    def show_prev_month(self):
        """Показывает предыдущий месяц"""
        self.current_month -= 1
        if self.current_month < 1:
            self.current_month = 12
            self.current_year -= 1
        self.update_calendar()

    def show_next_month(self):
        """Показывает следующий месяц"""
        self.current_month += 1
        if self.current_month > 12:
            self.current_month = 1
            self.current_year += 1
        self.update_calendar()

    def on_calendar_click(self, event):
        """Корректный обработчик клика по календарю"""
        # Получаем размеры холста
        canvas_width = self.calendar_canvas.winfo_width()
        canvas_height = self.calendar_canvas.winfo_height()

        # Рассчитываем параметры отображения
        first_day = datetime(self.current_year, self.current_month, 1).date()
        start_weekday = first_day.weekday()  # 0-пн, 6-вс

        # Корректное определение последнего дня месяца
        if self.current_month == 12:
            last_day = datetime(self.current_year + 1, 1, 1).date() - timedelta(days=1)
        else:
            last_day = datetime(self.current_year, self.current_month + 1, 1).date() - timedelta(days=1)

        total_days = last_day.day
        num_weeks = 6 if (start_weekday + total_days) > 35 else 5

        # Вычисляем размеры ячеек
        header_height = 30  # Высота строки с днями недели
        day_width = canvas_width / 7
        day_height = (canvas_height - header_height) / num_weeks

        # Определяем координаты клика
        click_y = event.y - header_height  # Корректируем с учетом заголовка

        # Пропускаем клики в заголовке
        if click_y < 0:
            return

        # Определяем строку и столбец
        day_col = int(event.x // day_width)
        day_row = int(click_y // day_height)

        # Пропускаем клики вне сетки дней
        if day_col < 0 or day_col > 6 or day_row < 0:
            return

        # Вычисляем номер дня
        day_num = day_row * 7 + day_col - start_weekday + 1

        # Проверяем, что день существует в этом месяце
        if day_num < 1 or day_num > total_days:
            return

        # Получаем корректную дату
        current_date = datetime(self.current_year, self.current_month, day_num).date()

        # Получаем заказы на эту дату
        orders = get_production_orders()
        day_orders = [o for o in orders
                      if datetime.strptime(o[3], "%Y-%m-%d").date() == current_date]

        if not day_orders:
            messagebox.showinfo("Информация", f"На {current_date.strftime('%d.%m.%Y')} нет заказов")
            return

        # Показываем диалог выбора действия
        self.show_date_actions(current_date)

    def show_date_actions(self, date):
        """Диалог выбора действия для даты"""
        dialog = tk.Toplevel(self)
        dialog.title(f"Действия для {date.strftime('%d.%m.%Y')}")
        dialog.geometry("300x150")
        dialog.grab_set()  # Модальное окно

        CTkLabel(dialog, text="Что вы хотите просмотреть?").pack(pady=10)

        btn_frame = CTkFrame(dialog)
        btn_frame.pack(pady=10)

        # Кнопки выбора
        CTkButton(btn_frame, text="Заказы",
                  command=lambda: [self.show_orders_for_date(date), dialog.destroy()]).pack(side=tk.LEFT, padx=5)

        CTkButton(btn_frame, text="Материалы",
                  command=lambda: [self.show_materials_for_date(date), dialog.destroy()]).pack(side=tk.LEFT, padx=5)


    def show_materials_for_date(self, date):
        """Показывает таблицу материалов для выбранной даты"""
        dialog = tk.Toplevel(self)
        dialog.title(f"Материалы на {date.strftime('%d.%m.%Y')}")
        dialog.geometry("800x600")

        # Создаем фрейм для таблицы
        frame = CTkFrame(dialog)
        frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Создаем таблицу
        columns = ("material", "required", "in_stock")
        tree = ttk.Treeview(frame, columns=columns, show="headings")

        # Настраиваем столбцы
        tree.heading("material", text="Материал")
        tree.heading("required", text="Требуется")
        tree.heading("in_stock", text="На складе")

        tree.column("material", width=400)
        tree.column("required", width=150, anchor='center')
        tree.column("in_stock", width=150, anchor='center')

        # Настраиваем тег для недостающих материалов
        tree.tag_configure('not_enough', background='#ffcccc')

        # Получаем все заказы на эту дату
        orders = [o for o in get_production_orders()
                  if datetime.strptime(o[3], "%Y-%m-%d").date() == date]

        # Собираем все материалы для этих заказов
        materials_sum = {}
        for order in orders:
            materials = get_materials_for_production_order(order[0])
            for mat in materials:
                mat_type = mat[1]
                amount = mat[2]
                dimension = mat[3]

                if mat_type not in materials_sum:
                    materials_sum[mat_type] = {
                        'amount': 0.0,
                        'dimension': dimension
                    }
                materials_sum[mat_type]['amount'] += amount

        # Функция поиска материала на складе
        def find_material_on_warehouse(material_name):
            warehouses = [
                get_all_component_materials(),
                get_all_film_materials(),
                get_all_window_materials(),
                get_all_triplex_materials(),
                get_all_main_glass_materials(),
                get_all_cutting_materials()
            ]

            for warehouse in warehouses:
                for item in warehouse:
                    if material_name.lower() in item[2].lower():
                        balance = item[8] if len(item) > 8 else item[7]
                        unit = item[3]
                        try:
                            return {
                                'amount': float(balance) if balance else 0.0,
                                'unit': unit
                            }
                        except (ValueError, TypeError):
                            return {'amount': 0.0, 'unit': 'шт.'}
            return None

        # Заполняем таблицу
        for mat_type, data in sorted(materials_sum.items()):
            required = f"{round(data['amount'], 2)} {data['dimension']}"
            warehouse = find_material_on_warehouse(mat_type)

            if warehouse:
                in_stock = f"{round(warehouse['amount'], 2)} {warehouse['unit']}"
                # Проверяем, хватает ли материала
                try:
                    if float(data['amount']) > float(warehouse['amount']):
                        tags = ('not_enough',)
                    else:
                        tags = ()
                except ValueError:
                    tags = ()
            else:
                in_stock = "Не найден"
                tags = ('not_enough',)

            tree.insert("", tk.END, values=(mat_type, required, in_stock), tags=tags)

        # Добавляем скроллбар
        scrollbar = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)

        # Упаковываем элементы
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # Кнопка закрытия
        CTkButton(dialog, text="Закрыть", command=dialog.destroy).pack(pady=10)

    def show_orders_for_date(self, date):
        """Показывает список заказов на выбранную дату с правильными столбцами"""
        dialog = tk.Toplevel(self)
        dialog.title(f"Заказы на {date.strftime('%d.%m.%Y')}")
        dialog.geometry("900x600")

        # Создаем фрейм для таблицы
        frame = CTkFrame(dialog)
        frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Создаем таблицу заказов с правильными столбцами
        columns = ("id", "name", "customer", "deadline", "priority", "status")
        tree = ttk.Treeview(frame, columns=columns, show="headings")

        # Настраиваем столбцы
        tree.heading("id", text="ID")
        tree.heading("name", text="Название заказа")
        tree.heading("customer", text="Заказчик")
        tree.heading("deadline", text="Срок выполнения")
        tree.heading("priority", text="Приоритет")
        tree.heading("status", text="Статус")

        tree.column("id", width=50, anchor='center')
        tree.column("name", width=250)
        tree.column("customer", width=200)
        tree.column("deadline", width=100, anchor='center')
        tree.column("priority", width=100, anchor='center')
        tree.column("status", width=100, anchor='center')

        # Получаем заказы на эту дату
        orders = [o for o in get_production_orders()
                  if datetime.strptime(o[3], "%Y-%m-%d").date() == date]

        # Заполняем таблицу с правильным порядком данных
        for order in orders:
            # Форматируем дату для отображения
            deadline = datetime.strptime(order[3], "%Y-%m-%d").strftime("%d.%m.%Y")
            tree.insert("", tk.END, values=(
                order[0],  # ID
                order[1],  # Название
                order[2],  # Заказчик
                deadline,  # Срок выполнения
                order[4],  # Приоритет
                order[5]  # Статус
            ))

        # Добавляем скроллбар
        scrollbar = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)

        # Упаковываем элементы
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # Кнопка для просмотра деталей заказа
        def view_order_details():
            selected = tree.focus()
            if not selected:
                return
            order_id = tree.item(selected)['values'][0]
            self.show_order_details_by_id(order_id)
            dialog.destroy()

        # Кнопки управления
        btn_frame = CTkFrame(dialog)
        btn_frame.pack(pady=10)

        CTkButton(btn_frame, text="Просмотреть детали", command=view_order_details).pack(side=tk.LEFT, padx=5)
        CTkButton(btn_frame, text="Закрыть", command=dialog.destroy).pack(side=tk.RIGHT, padx=5)

    def show_order_details_by_id(self, order_id):
        """Показывает детали заказа по ID"""
        self.current_order_id = order_id
        orders = get_production_orders()

        for order in orders:
            if order[0] == order_id:
                order_id, name, customer, deadline, priority, status = order
                deadline_date = datetime.strptime(deadline, "%Y-%m-%d").strftime("%d.%m.%Y")

                details = f"Заказ №{order_id}\n\n"
                details += f"Название: {name}\n"
                details += f"Заказчик: {customer}\n"
                details += f"Срок выполнения: {deadline_date}\n"
                details += f"Приоритет: {priority}\n"
                details += f"Статус: {status}\n"

                self.order_details_text.config(state=tk.NORMAL)
                self.order_details_text.delete(1.0, tk.END)
                self.order_details_text.insert(tk.END, details)
                self.order_details_text.config(state=tk.DISABLED)

                # Загружаем стеклопакеты и материалы для этого заказа
                self.load_windows_for_order(order_id)
                self.load_materials_for_order(order_id)

                # Обновляем данные склада для заказов на эту дату
                self.load_warehouse_data()
                break


    def add_production_order(self):
        """Добавление нового производственного заказа"""
        name = self.order_name_entry.get()
        customer = self.customer_entry.get()
        deadline = self.deadline_entry.get()
        priority = self.priority_var.get()

        if not name or not deadline:
            messagebox.showerror("Ошибка", "Название и срок обязательны для заполнения")
            return

        try:
            deadline_date = datetime.strptime(deadline, "%d.%m.%y").date()
        except ValueError:
            try:
                deadline_date = datetime.strptime(deadline, "%d.%m.%Y").date()
            except ValueError:
                messagebox.showerror("Ошибка", "Неверный формат даты. Используйте ДД.ММ.ГГ или ДД.ММ.ГГГГ")
                return

        # Добавляем заказ в БД
        order_id = add_production_order(name, customer, deadline_date.strftime("%Y-%m-%d"), priority, "В ожидании")

        # Загружаем список заказов
        self.load_production_orders()

        # Устанавливаем текущий заказ
        self.current_order_id = order_id
        self.show_order_details_by_id(order_id)

        # Очищаем поля
        self.order_name_entry.delete(0, tk.END)
        self.customer_entry.delete(0, tk.END)
        self.deadline_entry.delete(0, tk.END)

        # Обновляем календарь
        self.update_calendar()

    def load_production_orders(self):
        """Загрузка и отображение заказов с группировкой по датам и раскрытием"""
        orders = get_production_orders()

        # Очистка старых виджетов
        for widget in self.orders_frame.winfo_children():
            widget.destroy()

        self.scrollable_frame = CTkScrollableFrame(self.orders_frame, fg_color="transparent")
        self.scrollable_frame.pack(fill=tk.BOTH, expand=True)

        from collections import defaultdict
        orders_by_date = defaultdict(list)

        for order in orders:
            deadline_date = datetime.strptime(order[3], "%Y-%m-%d").date()
            date_str = deadline_date.strftime("%d.%m.%Y")
            orders_by_date[date_str].append(order)

        self.order_frames = {}

        for date_str in sorted(orders_by_date.keys(), key=lambda d: datetime.strptime(d, "%d.%m.%Y"), reverse=True):
            order_list = orders_by_date[date_str]

            # Обёртка под дату и список заказов
            date_container = CTkFrame(self.scrollable_frame, fg_color="transparent")
            date_container.pack(fill=tk.X, pady=(0, 5))

            # Переменная: открыт ли список
            expanded_var = tk.BooleanVar(value=False)

            # Вложенный фрейм заказов
            orders_subframe = CTkFrame(date_container, fg_color="transparent")

            def toggle_orders(var, frame):
                if var.get():
                    frame.pack(fill=tk.X)
                else:
                    frame.pack_forget()

            # Создание замыкания с текущими переменными
            def make_toggle_callback(var, frame):
                return lambda: [var.set(not var.get()), toggle_orders(var, frame)]

            # Кнопка даты
            date_btn = CTkButton(
                date_container,
                text=f"{date_str} • {len(order_list)} заказ(ов)",
                command=make_toggle_callback(expanded_var, orders_subframe),
                anchor="w",
                fg_color="#3a3a3a",
                hover_color="#4a4a4a",
                corner_radius=5,
                font=("Arial", 12, "bold"),
                height=35
            )
            date_btn.pack(fill=tk.X)

            # Открыть, если текущий заказ в этой дате
            if any(str(o[0]) == str(self.current_order_id) for o in order_list):
                expanded_var.set(True)
                orders_subframe.pack(fill=tk.X)

            # Добавление заказов
            for order in sorted(order_list, key=lambda o: o[0]):
                order_id, name, _, _, _, status = order

                order_frame = CTkFrame(
                    orders_subframe,
                    fg_color="#1e88e5" if str(order_id) == str(self.current_order_id) else "#6e6e6e",
                    height=35,
                    corner_radius=5
                )
                order_frame.pack(fill=tk.X, padx=10, pady=2)

                label = CTkLabel(
                    order_frame,
                    text=f"{order_id}: {name} • {status}",
                    anchor="w",
                    text_color="white",
                    font=("Arial", 12)
                )
                label.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=10)

                def make_order_callback(oid):
                    return lambda e=None: self.select_order(oid)

                # Обработчики
                order_frame.bind("<Button-1>", make_order_callback(order_id))
                label.bind("<Button-1>", make_order_callback(order_id))

                self.order_frames[order_id] = order_frame

    def select_order(self, order_id):
        """Выбор и отображение заказа без перезагрузки списка"""
        # Убрать подсветку со всех
        for oid, frame in self.order_frames.items():
            frame.configure(fg_color="#1e88e5" if str(oid) == str(order_id) else "#6e6e6e")

        # Обновить текущий выбранный ID
        self.current_order_id = order_id

        # Показать детали заказа
        self.show_order_details_by_id(order_id)

    def show_order_details(self, event=None):
        """Отображение деталей выбранного заказа"""
        selection = self.orders_listbox.curselection()
        if not selection:
            return

        # Получаем ID заказа из строки (формат "ID: Название")
        order_text = self.orders_listbox.get(selection[0])
        order_id = order_text.split(":")[0].strip()

        self.current_order_id = order_id
        orders = get_production_orders()

        for order in orders:
            if str(order[0]) == order_id:
                order_id, name, customer, deadline, priority, status = order
                deadline_date = datetime.strptime(deadline, "%Y-%m-%d").strftime("%d.%m.%Y")

                details = f"Заказ №{order_id}\n\n"
                details += f"Название: {name}\n"
                details += f"Заказчик: {customer}\n"
                details += f"Срок выполнения: {deadline_date}\n"
                details += f"Приоритет: {priority}\n"
                details += f"Статус: {status}\n"

                # Обновляем текстовое поле
                self.order_details_text.config(state=tk.NORMAL)
                self.order_details_text.delete(1.0, tk.END)
                self.order_details_text.insert(tk.END, details)
                self.order_details_text.config(state=tk.DISABLED)

                # Загружаем связанные стеклопакеты
                self.load_windows_for_order(order_id)
                self.load_materials_for_order(order_id)

                self.load_warehouse_data()
                break

    def load_windows_for_order(self, order_id):
        """Загрузка списка стеклопакетов для заказа"""
        # Очищаем таблицу
        for item in self.windows_tree.get_children():
            self.windows_tree.delete(item)

        # Загружаем данные из БД
        windows = get_windows_for_production_order(order_id)

        # Добавляем с порядковым номером
        for i, window in enumerate(windows, 1):
            self.windows_tree.insert("", tk.END, values=(i, *window[1:]))  # Пропускаем id и order_id

    def add_window_to_order(self):
        """Добавление стеклопакета к заказу"""
        if not self.current_order_id:
            messagebox.showwarning("Предупреждение", "Сначала выберите заказ")
            return

        # Создаем диалоговое окно
        self.window_dialog = tk.Toplevel(self)
        self.window_dialog.title("Добавить стеклопакет")
        self.window_dialog.geometry("300x250")
        self.window_dialog.protocol("WM_DELETE_WINDOW", self.close_window_dialog)

        # Переменные для хранения значений
        self.width_var = tk.StringVar()
        self.height_var = tk.StringVar()
        self.quantity_var = tk.StringVar()
        self.type_var = tk.StringVar()

        # Элементы управления
        CTkLabel(self.window_dialog, text="Ширина (мм):").pack(pady=5)
        width_entry = CTkEntry(self.window_dialog, textvariable=self.width_var)
        width_entry.pack(pady=5)
        width_entry.focus_set()

        CTkLabel(self.window_dialog, text="Высота (мм):").pack(pady=5)
        height_entry = CTkEntry(self.window_dialog, textvariable=self.height_var)
        height_entry.pack(pady=5)

        CTkLabel(self.window_dialog, text="Количество:").pack(pady=5)
        quantity_entry = CTkEntry(self.window_dialog, textvariable=self.quantity_var)
        quantity_entry.pack(pady=5)

        CTkLabel(self.window_dialog, text="Тип стеклопакета:").pack(pady=5)
        type_entry = CTkEntry(self.window_dialog, textvariable=self.type_var)
        type_entry.pack(pady=5)


        # Кнопки
        button_frame = CTkFrame(self.window_dialog)
        button_frame.pack(pady=10)

        CTkButton(button_frame, text="Сохранить и добавить еще",
                  command=self.save_window_and_continue).pack(side=tk.LEFT, padx=5)
        CTkButton(button_frame, text="Сохранить и закрыть",
                  command=self.save_window_and_close).pack(side=tk.LEFT, padx=5)
        CTkButton(button_frame, text="Отмена",
                  command=self.close_window_dialog).pack(side=tk.LEFT, padx=5)

    def save_window_and_continue(self):
        """Сохраняет текущий стеклопакет и очищает поля для ввода следующего"""
        try:
            width = int(self.width_var.get())
            height = int(self.height_var.get())
            quantity = int(self.quantity_var.get())
            window_type = self.type_var.get()

            if width <= 0 or height <= 0:
                messagebox.showerror("Ошибка", "Размеры должны быть положительными числами")
                return

            if quantity <= 0:
                messagebox.showerror("Ошибка", "Количество должно быть положительным")
                return

            add_window_to_production_order(self.current_order_id, window_type, width, height, quantity)
            self.load_windows_for_order(self.current_order_id)

            # Очищаем поля для ввода следующего значения
            self.width_var.set("")
            self.height_var.set("")
            self.quantity_var.set("")
            self.window_dialog.focus_set()  # Возвращаем фокус в окно

        except ValueError:
            messagebox.showerror("Ошибка", "Введите корректные числовые значения")

    def save_window_and_close(self):
        """Сохраняет стеклопакет и закрывает окно"""
        self.save_window_and_continue()
        self.close_window_dialog()

    def close_window_dialog(self):
        """Закрывает окно добавления стеклопакетов"""
        if hasattr(self, 'window_dialog') and self.window_dialog:
            self.window_dialog.destroy()

    def delete_window_from_order(self):
        """Удаление стеклопакета из заказа"""
        if not self.current_order_id:
            messagebox.showwarning("Предупреждение", "Сначала выберите заказ")
            return

        selection = self.windows_tree.selection()
        if not selection:
            messagebox.showwarning("Предупреждение", "Выберите стеклопакет для удаления")
            return

        if messagebox.askyesno("Подтверждение", "Удалить выбранный стеклопакет?"):
            # print(selection, "===win==", selection[0])
            window_id = self.windows_tree.item(selection[0], "values")[0]
            delete_window_from_production_order(window_id)
            self.load_windows_for_order(self.current_order_id)

    def load_warehouse_data(self):
        """Автоматическая загрузка данных склада при выборе заказа"""
        # Очищаем таблицу
        for item in self.warehouse_tree.get_children():
            self.warehouse_tree.delete(item)

        # Если нет выбранного заказа - показываем пустую таблицу
        if not self.current_order_id:
            return

        # Получаем информацию о текущем заказе
        orders = get_production_orders()
        current_order = None
        for order in orders:
            if order[0] == self.current_order_id:
                current_order = order
                break

        if not current_order:
            return

        # Получаем дату текущего заказа
        try:
            order_date = datetime.strptime(current_order[3], "%Y-%m-%d").date()
            date_str = order_date.strftime("%d.%m.%Y")
        except:
            return

        # Обновляем текст вкладки через Notebook
        tab_index = self.details_notebook.index(self.warehouse_tab)
        self.details_notebook.tab(tab_index, text=f"Склад (на {date_str})")

        # Получаем все заказы на эту же дату
        orders_on_date = [o for o in get_production_orders()
                          if datetime.strptime(o[3], "%Y-%m-%d").date() == order_date]

        # Создаем словарь для хранения суммарных количеств материалов
        materials_sum = {}

        # Собираем данные по всем заказам на эту дату
        for order in orders_on_date:
            order_id = order[0]
            materials = get_materials_for_production_order(order_id)

            for material in materials:
                mat_type = material[1]
                amount = material[2]
                dimension = material[3]

                if mat_type not in materials_sum:
                    materials_sum[mat_type] = {
                        'amount': 0.0,
                        'dimension': dimension
                    }

                materials_sum[mat_type]['amount'] += amount

        # Функция для поиска материала на складах
        def find_material_on_warehouse(material_name):
            warehouses = [
                get_all_component_materials(),
                get_all_film_materials(),
                get_all_window_materials(),
                get_all_triplex_materials(),
                get_all_main_glass_materials(),
                get_all_cutting_materials()
            ]

            for warehouse in warehouses:
                for item in warehouse:
                    warehouse_name = item[2]
                    if material_name.lower() in warehouse_name.lower():
                        balance = item[8] if len(item) > 8 else item[7]
                        unit = item[3]
                        try:
                            return {
                                'amount': float(balance) if balance else 0.0,
                                'unit': unit
                            }
                        except (ValueError, TypeError):
                            return {'amount': 0.0, 'unit': 'шт.'}
            return None

        # Сортируем материалы по алфавиту
        sorted_materials = sorted(materials_sum.items(), key=lambda x: x[0])

        # Добавляем данные в таблицу
        for mat_type, data in sorted_materials:
            used_value = f"{round(data['amount'], 3)} {data['dimension']}"
            warehouse_material = find_material_on_warehouse(mat_type)

            # Определяем стиль строки (красный если материала не хватает)
            tags = ()
            if warehouse_material:
                try:
                    if float(data['amount']) > float(warehouse_material['amount']):
                        tags = ('not_enough',)
                except ValueError:
                    pass

            stock_value = f"{round(warehouse_material['amount'], 3)} {warehouse_material['unit']}" if warehouse_material else "Не найден"

            self.warehouse_tree.insert("", tk.END,
                                       values=(mat_type, used_value, stock_value),
                                       tags=tags)

        # Настраиваем стиль для недостающих материалов
        self.warehouse_tree.tag_configure('not_enough', background='#ffcccc')

    def load_materials_for_order(self, order_id):
        """Загрузка списка материалов для заказа с нумерацией"""
        # Очищаем таблицу
        for item in self.materials_tree.get_children():
            self.materials_tree.delete(item)

        # Загружаем данные из БД
        materials = get_materials_for_production_order(order_id)

        # Добавляем материалы в таблицу с порядковым номером
        for i, material in enumerate(materials, 1):
            self.materials_tree.insert("", tk.END, values=(i, *material[1:]))

    def add_material_to_order(self):
        """Добавление материала к заказу"""
        if not self.current_order_id:
            messagebox.showwarning("Предупреждение", "Сначала выберите заказ")
            return

        # Создаем диалоговое окно
        self.material_dialog = tk.Toplevel(self)
        self.material_dialog.title("Добавить материал")
        self.material_dialog.geometry("300x250")
        self.material_dialog.protocol("WM_DELETE_WINDOW", self.close_window_dialog)

        # Переменные для хранения значений
        self.material_type_var = tk.StringVar()
        self.material_amount_var = tk.StringVar(value="1.0")
        self.material_dimension_var = tk.StringVar(value="шт.")

        # Элементы управления
        CTkLabel(self.material_dialog, text="Тип материала:").pack(pady=5)
        type_entry = CTkEntry(self.material_dialog, textvariable=self.material_type_var)
        type_entry.pack(pady=5)
        type_entry.focus_set()

        CTkLabel(self.material_dialog, text="Количество:").pack(pady=5)
        amount_entry = CTkEntry(self.material_dialog, textvariable=self.material_amount_var)
        amount_entry.pack(pady=5)

        CTkLabel(self.material_dialog, text="Единица измерения:").pack(pady=5)
        dimension_entry = CTkEntry(self.material_dialog, textvariable=self.material_dimension_var)
        dimension_entry.pack(pady=5)

        # Кнопки
        button_frame = CTkFrame(self.material_dialog)
        button_frame.pack(pady=10)

        CTkButton(button_frame, text="Сохранить",
                  command=self.save_material).pack(side=tk.LEFT, padx=5)
        CTkButton(button_frame, text="Отмена",
                  command=self.material_dialog.destroy).pack(side=tk.RIGHT, padx=5)
        self.load_warehouse_data()

    def save_material(self):
        """Сохранение материала в БД"""
        try:
            material_type = self.material_type_var.get()
            amount = float(self.material_amount_var.get())
            dimension = self.material_dimension_var.get()

            if not material_type or not dimension:
                messagebox.showerror("Ошибка", "Все поля должны быть заполнены")
                return

            if amount <= 0:
                messagebox.showerror("Ошибка", "Количество должно быть положительным")
                return

            # Добавляем материал в БД
            add_material_to_production_order(
                self.current_order_id,
                material_type,
                amount,
                dimension
            )

            # Обновляем таблицу материалов
            self.load_materials_for_order(self.current_order_id)

            # Закрываем диалоговое окно
            self.material_dialog.destroy()

        except ValueError:
            messagebox.showerror("Ошибка", "Введите корректные числовые значения")

    def delete_material_from_order(self):
        """Удаление материала из заказа"""
        if not self.current_order_id:
            messagebox.showwarning("Предупреждение", "Сначала выберите заказ")
            return

        selection = self.materials_tree.selection()
        if not selection:
            messagebox.showwarning("Предупреждение", "Выберите материал для удаления")
            return

        if messagebox.askyesno("Подтверждение", "Удалить выбранный материал?"):
            # print(selection, "===mat==", selection[0])
            material_id = self.materials_tree.item(selection[0], "values")[0]
            delete_material_from_production_order(material_id)
            self.load_materials_for_order(self.current_order_id)
            self.load_warehouse_data()

    def change_order_status(self, new_status):
        """Изменение статуса заказа"""
        if not self.current_order_id:
            messagebox.showwarning("Предупреждение", "Выберите заказ")
            return

        update_production_order_status(self.current_order_id, new_status)
        self.load_production_orders()

        if self.current_order_id:
            self.show_order_details_by_id(self.current_order_id)

        self.update_calendar()

    def delete_order(self):
        """Удаление производственного заказа"""
        if not self.current_order_id:
            messagebox.showwarning("Предупреждение", "Выберите заказ")
            return

        if messagebox.askyesno("Подтверждение", "Удалить выбранный заказ?"):
            delete_production_order(self.current_order_id)
            self.load_production_orders()

            # Очищаем детали заказа
            self.order_details_text.config(state=tk.NORMAL)
            self.order_details_text.delete(1.0, tk.END)
            self.order_details_text.config(state=tk.DISABLED)

            # Очищаем таблицы
            for item in self.windows_tree.get_children():
                self.windows_tree.delete(item)

            for item in self.materials_tree.get_children():
                self.materials_tree.delete(item)

            self.current_order_id = None
            self.update_calendar()
            self.load_warehouse_data()

    def import_order_from_excel(self):
        """Импорт заказа из Excel (всегда создаёт новый заказ)"""
        file_path = filedialog.askopenfilename(
            title="Выберите файл заказа",
            filetypes=[("Excel files", "*.xls *.xlsx"), ("All files", "*.*")]
        )
        if not file_path:
            return

        try:
            order_data = parse_excel_order(file_path)

            # 2. Создаём НОВЫЙ заказ (игнорируем self.current_order_id)
            deadline = datetime.strptime(order_data['deadline'], "%d.%m.%Y").strftime("%Y-%m-%d")
            new_order_id = add_production_order(
                name=order_data['order_name'],
                customer=order_data['customer'],
                deadline=deadline,
                priority="Средний",
                status="В ожидании"
            )

            # 3. Добавляем все стеклопакеты из файла в новый заказ
            for window in order_data['windows']:
                add_window_to_production_order(
                    order_id=new_order_id,
                    window_type=window['type'],
                    width=window['width'],
                    height=window['height'],
                    quantity=window['quantity']
                )

            # 4. Добавляем все материалы из файла в новый заказ
            for material in order_data['materials']:
                add_material_to_production_order(
                    order_id=new_order_id,
                    material_type=material['type'],
                    amount=material['amount'],
                    dimension=material['dimension']
                )

            # 5. Обновляем интерфейс
            self.load_production_orders()
            self.load_warehouse_data()
            self.current_order_id = new_order_id
            self.show_order_details_by_id(new_order_id)
            messagebox.showinfo("Успех", "Новый заказ создан и данные импортированы!")

        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка импорта:\n{str(e)}")
            print(f"DEBUG: {e}")


